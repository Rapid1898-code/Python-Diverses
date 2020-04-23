import requests
import csv
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import random
import subprocess
import time
import re

# Ausgabe der Liste als XLS-File inkl. Prüfung ob Datei geöffnet ist
# Input stock: Name der Aktie
# Input content: Inhalt in Listenform
# Input filenmae: Name des XLSX-File
# Input append: 1=>anhängen von neuen Worksheets, 0=>überschreiben des XLS
def save_xls(stock, content, filename, append):
    #check ob append ausgewählt - aber file nicht vorhanden - dann Wechsel über Überschreibmodus 0
    try:
        book = load_workbook (filename)
    except:
        append = 0
    while True:
        try:
            if append == 0:
                writer = pd.ExcelWriter(filename, engine = 'openpyxl', options={'strings_to_numbers': True})
            elif append == 1:
                book = load_workbook (filename)
                writer = pd.ExcelWriter(filename, engine = 'openpyxl', options={'strings_to_numbers': True})
                writer.book = book
            pd.DataFrame(content).to_excel (writer, sheet_name=stock, header=False, index=False)
            writer.save ()
            writer.close ()
            break
        except Exception as e:
            print(e)
            input ("Datei ist evt. bereits geöffnet - bitte schließen und <Enter> drücken!")


def read_stamm(stock):
    output = []
    link = "https://www.ariva.de" + stock + "/bilanz-guv?page=" + "0" + "#stammdaten"
    page = requests.get (link)
    soup = BeautifulSoup (page.content, "html.parser")

    # Stammdaten auslesen
    output.append(["Stammdaten",""])
    table = soup.find_all ("div", class_="column half")
    for i in table:
        for j in i.find_all ("tr"):
            row = []
            for k in j.find_all ("td"):
                if k.text.strip() == "": row.append ("-")
                else: row.append (k.text.strip())
            output.append (row)

    # Kontakte auslesen
    output[0].extend(["Kontakt",""])
    table = soup.find_all ("div", class_="column half last")
    nr = 1
    for i in table:
        for j in i.find_all ("tr"):
            for k in j.find_all ("td"):
                if k.text.strip() == "": output[nr].append ("-")
                else: output[nr].append (k.text.strip())
            nr += 1

    # Termine auslesen
    table = soup.find_all ("div", class_="termine abstand new")
    nr = 1
    # Termine - Überschrift auslesen
    for i in table:
        cont = i.find("h3", class_="arhead undef").text.strip()
        output[0].append(cont)
    # Termine Inhalt auslesen
    for i in table:
        for j in i.find_all ("tr"):
            for k in j.find_all ("td"):
                #Prüfung ob Tag.Monat damit Jahr ergänzt wird
                pattern = '^[0-9][0-9].[0-9][0-9].$'
                if re.match(pattern, k.text.strip()): output[nr].append (k.text.strip()+cont[-4:])
                else: output[nr].append (k.text.strip())
            nr += 1

    # Aktionäre
    output[0].extend(["Aktionäre",""])
    table = soup.find_all ("div", class_="aktStruktur abstand new")
    nr = 1
    for i in table:
        for j in i.find_all ("tr"):
            for k in j.find_all ("td"):
                output[nr].append (k.text.strip())
            nr += 1

    # Management / Aufsichtsrat
    table = soup.find_all ("div", class_="management abstand new")
    for i in table:
        for j in i.find_all ("tr"):
            row = []
            for k in j.find_all ("td"):
                row.append (k.text.strip())
            output.append (row)

    # Profil
    table  = soup.find(id="profil_text")
    txt = ""
    for i in table.find_all ("p"):
        if txt == "": txt = txt + i.text.strip()
        else: txt = txt + " " + i.text.strip()
    output.append(["Profil",txt])

    return(output)

output = read_stamm("/apple-aktie")
save_xls("Apple",output,"Test2.xlsx",0)


