from openpyxl import load_workbook
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side

l1 = [["1","2","3"],["23897492874982719812798127","5","6"],["7345243523","8","12345"]]

def save_xls(stock, content, filename):
    column_widths = []
    for row in l1:
        for i, cell in enumerate (row):
            if len (column_widths) > i:
                if len (cell) > column_widths[i]:
                    column_widths[i] = len (cell)
            else:
                column_widths += [len (cell)]

    writer = pd.ExcelWriter(filename, engine = 'openpyxl')
    pd.DataFrame(content).to_excel (writer, sheet_name=stock, header=False, index=False)
    ws = writer.sheets[stock]


    for i, column_width in enumerate (column_widths):
        ws.column_dimensions[get_column_letter (i + 1)].width = column_width+1

    #headline = Font(bold=True) and PatternFill(fill_type="solid", start_color='fcba03',end_color='fcba03')
    bold = Font(bold=True)
    bg_yellow = PatternFill(fill_type="solid", start_color='fcba03',end_color='fcba03')
    bg_green = PatternFill(fill_type="solid", start_color='c7ffcd',end_color='fffbc7')
    fr = Border (outline=Side (style='thin'),right=Side (style='thin'),top=Side (style='thin'),bottom=Side (style='thin'))
    #fr = Border(outline=Side(border_style="thin",color='FF000000'))

    for cell in ws["1:1"]:
        cell.font = bold
        cell.fill = bg_yellow
        cell.border = fr
    ws["C3"].font = bold
    ws["B2"].fill = bg_green

    for cell in ws["A1:A2"]: cell[0].fill = bg_green

    writer.save ()
    writer.close ()

"""
    a1=writer.sheets[stock]["A1:B1"]
    a1.font = bold
    a1.fill = bg_yellow
"""





save_xls("test", l1, "TEMP3.xlsx")
