import xlrd
from datetime import datetime
from openpyxl import load_workbook
import pandas as pd

# read input-excel-sheets
filename = "Prices_Apple.xlsx"
#old wb_price = xlrd.open_workbook(filename)
#old sh_price = wb_price.sheet_by_index(0)
wb_price = load_workbook(filename)
sh_price  = wb_price.active
stock = wb_price.sheetnames[0]
wb_data = load_workbook("Data_Apple.xlsx")
sh_data = wb_data.active

# read price xls in list
price_list = []
for row in sh_price.iter_rows():
    zeile = []
    for cell in row:
        if cell.value is None: zeile.append("")
        else: zeile.append(cell.value)
    price_list.append(zeile)
print(price_list)

# read data xls in list
data_list = []
for row in sh_data.iter_rows():
    zeile = []
    for cell in row:
        if cell.value is None: zeile.append("")
        else: zeile.append(cell.value)
    data_list.append(zeile)
print(data_list)

# find necessary rows
row_title = row_shares = 0
for i in range(len(data_list)-1):
    if "Bilanz in Mio." in data_list[i][0]: row_title = data_list[i]
    if "Aktien im Umlauf"  in data_list[i][0]: row_shares= data_list[i]

# caculate marketcap per day with daily price and yearly outstanding shares
tmp_year=0
tmp_shares=0
for i in range (1,len(price_list)):
    if datetime.strptime(price_list[i][0],"%d.%m.%Y") != tmp_year:
        for j in range(2,len(row_title)):
            if datetime.strptime(price_list[i][0],"%d.%m.%Y").year-1 == int(row_title[j]):
                tmp_year = int(row_title[j])
                tmp_shares = row_shares[j]
                break
    price_list[i][2]  = round(price_list[i][1] * tmp_shares / 1000,2)

# Überschrift ergänzen
price_list.insert(1,["","Price","MarketCap"])
price_list.insert(2,["","Kurs","MarktKap"])

#print (data_list)
#print (price_list)
#print(row_title)
#print(row_shares)

#  save XLSX
book  = load_workbook (filename)
writer = pd.ExcelWriter (filename, engine='openpyxl', options={'strings_to_numbers': True})
writer.book = book
pd.DataFrame (price_list).to_excel (writer, sheet_name=stock, header=False, index=False)
while True:
    try:
        writer.save ()
        writer.close ()
        break
    except Exception as e:
        print ("Error: ", e)
        input ("Datei kann nicht geöffnet werden - bitte schließen und <Enter> drücken!")
