import xlrd     # Einlesen von Excel-Sheet
import requests
import csv
from datetime import datetime
import datetime
import calendar
from bs4 import BeautifulSoup
import pandas as pd
from itertools import zip_longest
import timeit
import time
import random
import subprocess
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side

#TODO
# Marktkapitalisierung pro Tag im Kurs-XLS ergänzen (evt. mit eigenem Programm)

# Unternehmen eines bestimmten Index werden eingelesen
# Output: Dict in der Form Kürzel von Ariva.de + Name des Titels (z.b '/apple-aktie': 'Apple')
def read_index(index_name):
    page_nr=0
    index_stocks = {}
    temp_stocks = {}
    while True:
        page = requests.get ("https://www.ariva.de/"+index_name+"?page="+str(page_nr))
        soup = BeautifulSoup (page.content, "html.parser")
        table  = soup.find(id="result_table_0")
        for row  in table.find_all("td"):
            if row.get("class") == ["ellipsis", "nobr", "new", "padding-right-5"]:
                index_stocks[row.find("a")["href"][1:]] = row.text.strip().capitalize()
        #Dict sortieren nach Value
        index_stocks = {k: v for k, v in sorted(index_stocks.items(), key=lambda item: item[1])}
        if temp_stocks == index_stocks: break
        page_nr += 1
        temp_stocks = index_stocks
    return(index_stocks)

# VPN-Switch bei NordVPN mit x Sekunden Verzögerung
# Output: Rückgabe des zufällig gewählten Landes
def vpn_switch(sek):
    countries = ["Austria", "Belgium", "Germany", "Italy", "Poland", "Romania", "Serbia", "Switzerland", "United Kingdom",
                 "Croatia","Estonia","France","Greece","Latvia","Netherlands"]
    rand_country = random.randrange(len(countries)-1)
    subprocess.call (["C:/Program Files (x86)/NordVPN/NordVPN.exe", "-c", "-g", countries[rand_country]])
    print ("VPN Switch to",countries[rand_country],"...")
    for i in range (sek, 0, -1):
        sys.stdout.write (str (i) + ' ')
        sys.stdout.flush ()
        time.sleep (1)
    print ("Connected to",countries[rand_country],"...")
    return(countries[rand_country])

# Aktienkurse für eine Unternehmen einlesen
# boerse_id ist fix =>  6 = Xetra
# Input Stock: Aktienkennung lt. Ariva.de z.b. /apple-aktie oder /wirecard-aktie
# Input Month: Monatsultimo im Format z.B. 2019-04-30
# Output: bei Spaltenposition 0: Ausgabe/Yield von Schlüsselwort "datum" + Inhalt des Datumsfeldes im Format dd.mm.yyyy
# Output: bei Spaltenposition 4: Ausgabe/Yield von Schlüsselwort "price" + Inhalt des Schlusskurses als Zahl
# Output: wenn IP-Seitensperre erfolgt ist (Kein Zugriff auf HTML-Seite im Text): Ausgabe/Yield von Schlüsselwort "abbruch" + Textinhalt
def stock_prices_month (stock,month,whg,boerse_id):
    # read table with monatlichen Kursen
    # Börse-ID: 6 => Xetra, 21 => NYSE, 40 => Nasdaq
    url = "https://www.ariva.de/" + stock + "/historische_kurse?boerse_id=" + str(boerse_id) + "&month=" + month + \
          "&currency=" + whg + "&clean_split=1&clean_split=0&clean_payout=0&clean_bezug=1&clean_bezug=0"
    page = requests.get (url)
    soup = BeautifulSoup (page.content, "html.parser")
    # Check ob Zugriff auf die IP-Adresse gesperrt ist
    # wenn ja wird "abbruch" zurückgegeben und dadurch die weitere Verarbeitung gestoppt
    if "Kein Zugriff" in soup.text:
        print("Abbruch!!!")
        yield "abbruch", soup.text.strip()
    for result in soup.find_all("tr", class_="arrow0"):
        for col_id, col_content in enumerate(result.find_all("td")):
            # wenn Ausgabejahr nicht mit dem gesuchten Jahr zusammenpasst => break
            if col_id == 0 and (datetime.datetime.strptime(col_content.text.strip(), "%d.%m.%y").year !=
                                datetime.datetime.strptime(month, "%Y-%m-%d").year or
                                datetime.datetime.strptime (col_content.text.strip (), "%d.%m.%y").month !=
                                datetime.datetime.strptime (month, "%Y-%m-%d").month):
                break
            # Datum des Kurse bzw. Schlusskurs auslesen
            if col_id == 0:     # 1.Spalte Datum
                yield "datum", col_content.text.strip()
            elif col_id == 4:   # 5.Spalte Schlusskurs
                yield "price", col_content.text.strip()

# Monatsultimo ermitteln für Zeitraum
# Input z.B. (3, 2015, datetime.now().month, datetime.now().year) bei Aufruf
# Output z.b. Ultimo-Datum z.b. 2016-04-30
def month_year_iter( start_month, start_year, end_month, end_year ):
    ym_start= 12*start_year + start_month - 1
    ym_end= 12*end_year + end_month - 1
    for ym in range(ym_end, ym_start-1, -1):
        y, m = divmod( ym, 12 )
        yield datetime.date(y,m+1,calendar.monthrange(y,m+1)[1])

# Historische Kurse eines Aktientitels lesen mit Angabe von Zeitraum
def read_prices(stock,start_month, start_year, end_month, end_year,whg):
    global abbruch
    output = [["Datum",stock.upper().replace("-"," ").replace("AKTIE","").replace("_"," "),"in "+whg]]
    year = datetime.datetime.now().year
    if year <= end_year: print(stock + " " + str(year))
    for i in month_year_iter(start_month, start_year, end_month, end_year):
        if abbruch == True: break
        #Ausgabe zur Fortschrittskontrolle des Programms
        if i.year != year:
            year -= 1
            if year <= end_year: print(stock + " " + str(year))

        temp_output = []
        for boerse_id in [6,21,40]:
            for j in stock_prices_month(stock,str(i),whg,boerse_id):
                if j[0] =="abbruch":
                    abbruch = True
                    break
                if j[0] == "datum":
                    temp_output.append(datetime.datetime.strptime(j[1],"%d.%m.%y").strftime("%d.%m.%Y"))
                elif j[0] == "price":
                    temp_output.append(float(j[1].replace(".","").replace(",",".")))
                elif j[0] == "blank": pass
        temp_month = []
        for k_id, k_cont in enumerate(temp_output):
            if k_id%2 == 0:
                if k_cont not in temp_month: temp_month.extend([temp_output[k_id], temp_output[k_id + 1]])
        temp_month2 = []
        for k in range(0, len(temp_month),2): temp_month2.append([temp_month[k],temp_month[k+1]])
        temp_month2.sort(reverse=True)
        for k in temp_month2: output.append(k)
    return  output

# Erstellung einer Datums-Liste vom aktuellstem bis zum  ältesten Datum im Format jjjj-mm-dd
def date_list (datum_von, datum_bis):
    mydates2  = []
    mydates = pd.date_range(datum_bis, datum_von).tolist()
    for i in range(len(mydates)-1,-1,-1): mydates2.append(mydates[i].strftime('%d.%m.%y'))
    return(mydates2)

# Ausgabe der Liste als CSV-File
def csv_write(result, filename):
    while True:
        try:
            with open (filename,"w",newline="") as fp:
                a = csv.writer(fp,delimiter=",")
                a.writerows(result)
                break
        except Exception as e:
            print ("Error: ", e)
            input ("Datei kann nicht geöffnet werden - bitte schließen und <Enter> drücken!")
    fp.close()

# Ausgabe der Liste als XLS-File inkl. Prüfung ob Datei geöffnet ist
# Input stock: Name der Aktie
# Input content: Inhalt in Listenform
# Input filenmae: Name des XLSX-File
# Input append: 1=>anhängen von neuen Worksheets, 0=>überschreiben des XLS
def save_xls(stock, content, filename):
    global writemodus

    #check ob append ausgewählt - aber wenn file nicht vorhanden - dann Wechsel über Überschreibmodus 0
    try:
        book = load_workbook (filename)
    except:
        writemodus = 0
    if writemodus == 0:
        writer = pd.ExcelWriter(filename, engine = 'openpyxl', options={'strings_to_numbers': True})
    elif writemodus == 1:
        book = load_workbook (filename)
        if stock in book.sheetnames:
            print ("Aktie: ",stock," bereits im XLS: ",filename," enthalten - Aktie wird übersprungen")
            return
        writer = pd.ExcelWriter(filename, engine = 'openpyxl', options={'strings_to_numbers': True})
        writer.book = book
    pd.DataFrame(content).to_excel (writer, sheet_name=stock, header=False, index=False)
    if writemodus == 0: writemodus = 1

    # Automatische Anpassung der Spalten nach best fit
    column_widths = []
    ws = writer.sheets[stock]
    # Ermittlung des längsten Wertes pro Spalte
    for row in content:
        for i, cell in enumerate (row):
            if len (column_widths) > i:
                if len (str (cell)) > column_widths[i]:
                    column_widths[i] = len (str (cell))
            else:
                column_widths += [len (str (cell))]
    for i, column_width in enumerate (column_widths):
        ws.column_dimensions[get_column_letter (i + 1)].width = column_width+2

    # Formatierung des Excel-Sheets
    bold = Font (bold=True)
    bg_yellow = PatternFill (fill_type="solid", start_color='fbfce1', end_color='fbfce1')
    bg_grey = PatternFill (fill_type="solid", start_color='babab6', end_color='babab6')
    bg_green = PatternFill (fill_type="solid", start_color='c7ffcd', end_color='fffbc7')
    frame_all = Border (left=Side (style='thin'), right=Side (style='thin'), top=Side (style='thin'),bottom=Side (style='thin'))
    frame_upanddown = Border (top=Side (style='thin'), bottom=Side (style='thin'))
    for cell in ws["A:A"]:
        cell.font = bold
        cell.fill = bg_green
        cell.border = frame_all
    for cell in ws["1:1"]:
        cell.font = bold
        cell.fill = bg_green
        cell.border = frame_all
    freeze = ws["B2"]
    ws.freeze_panes = freeze

    # Excel Sheet speichern
    while True:
        try:
            writer.save ()
            writer.close ()
            break
        except Exception as e:
            print ("Error: ", e)
            input ("Datei kann nicht geöffnet werden - bitte schließen und <Enter> drücken!")

#Check ob Aktie schon enthalten ist im XLS
def check_xls(stock,filename):
    global writemodus
    try:
        book = load_workbook (filename)
        if writemodus == 0:
            wahl = input("Es befinden sich Daten im Excel-Sheet "+filename+" wollen Sie wirklich die Datei überschreiben (j/n)=")
            if wahl.upper() == "N": writemodus = 1
    except:
        writemodus = 0
        return False
    if stock in book.sheetnames and writemodus == 1:
        print ("Aktie: ",stock," bereits im XLS: ",filename," enthalten - Aktie wird übersprungen")
        return True

# Spalten bereinigen der Tabelle

# Datum vereinheitlichen + leere Spalten löschen
def spalten_bereinigen(output):
    # Datum-Spalten vereinheitlichen
    for i in range (2, len (output[0]) - 1):
        for j in range (1, len (output) - 1):
            if j % 2 == 1:
                if i == len (output[j]):
                    output[j].insert (i, "")
                    output[j + 1].insert (i, "")
                elif output[j][i] != output[0][i]:
                    output[j].insert (i, "")
                    output[j + 1].insert (i, "")
            else:
                continue
    # Check ob die letzte Spalte bei den Aktientiteln leer ist - sonst fehlt eine Spalte am Ende bei den Aktientiteln
    for i in range (1, len (output) - 1):
        if i % 2 == 1 and len (output[0]) != len (output[i]):
            output[i].insert (len (output[i]), "")
            output[i + 1].insert (len (output[i + 1]), "")

    # Leere Spalten löschen
    # kein Kurse für eine Aktie an diesem Tag
    pos_del = []
    for i in range (2, len (output[0])):
        empty = True
        for j in range (1, len (output) - 1):
            if output[j][i] != "":
                empty = False
                break
        if empty == True: pos_del.append (i)
    pos_del.reverse ()
    for k in pos_del:
        for m in range (len (output)):
            del output[m][k]
    return(output)

# Einlesen eines Ergebnis-XLS um die Datümer zusätzlich abzugleichen
# Durchgängiges Datum muss noch vorne ergänzt werden
def read_XLS():
    workbook = xlrd.open_workbook ("DAX Price2.xlsx")
    sheet = workbook.sheet_by_index (0)
    liste = []
    for row in range (sheet.nrows):
        zeile=[]
        for col in range(sheet.ncols):
            if sheet.cell(row,col).ctype == 3:
                zeile.append(datetime(*xlrd.xldate_as_tuple(sheet.cell(row,col).value,0)).strftime("%d.%m.%Y"))
            else:
                zeile.append(sheet.cell(row,col).value)
        liste.append(zeile)
    #print(liste)

    result = [list(filter(None,i)) for i in zip_longest(*liste)]
    print(result)
    csv_write(result, "testout.csv")

# DAX30 Unternehmen + Ex-Unternehmen
# Input: Für welche Aktientitel die Verarbeitung erfolgen soll
"""
stocks_dic = {'/apple-aktie': 'Apple', '/infineon-aktie': 'Infineon', '/volkswagen_vz-aktie': 'Volkswagen Vz', '/continental-aktie': 'Continental',
'/bmw-aktie': 'BMW St', '/heidelbergcement-aktie': 'HeidelbergCement', '/mtu_aero_engines-aktie': 'MTU Aero Engines', '/covestro-aktie': 'Covestro',
'/siemens-aktie': 'Siemens', '/daimler-aktie': 'Daimler', '/munich_re-aktie': 'Munich Re', '/basf-aktie': 'BASF', '/deutsche_bank-aktie': 'Dt. Bank',
'/deutsche_post-aktie': 'Dt. Post', '/adidas-aktie': 'adidas', '/allianz-aktie': 'Allianz', '/sap-aktie': 'SAP', '/deutsche_telekom-aktie': 'Dt. Telekom',
'/linde_plc-aktie': 'Linde PLC', '/bayer-aktie': 'Bayer', '/henkel_vz-aktie': 'Henkel Vz', '/lufthansa-aktie': 'Lufthansa', '/beiersdorf-aktie': 'Beiersdorf',
'/merck_kgaa-aktie': 'Merck KGaA', '/fresenius_medical_care-aktie': 'Fresenius Medical Care', '/deutsche_b%C3%B6rse-aktie': 'Dt. Börse', '/fresenius-aktie': 'Fresenius',
'/wirecard-aktie': 'Wirecard', '/rwe-aktie': 'RWE St', '/e-on-aktie': 'E.ON', '/vonovia-aktie': 'Vonovia', '/thyssenkrupp-akti': 'ThyssenKrupp',
'/commerzbank-aktie': 'Commerzbank', '/prosiebensat-1_media-aktie': 'ProSiebenSat-1 Media', '/uniper-aktie': 'Uniper', '/k-s-6-aktie': 'KS6',
 '/lanxess-aktie': 'Lanxess', '/osram_licht-aktie': 'Osram Licht', '/ceconomy_st-aktie': 'Ceconomy St', '/man-aktie': 'MAN', '/salzgitter-aktie': 'Salzgitter'
 , '/hannover_rück-aktie': 'Hannover Rück', '/tui-aktie': 'TUI', '/mlp-aktie': 'MLP'}
"""

stocks_dic = {'apple-aktie': 'Apple'}

#Input-Parameter
#Input - Angabe welcher Index gelesen werden soll (z.B. DAX-30) - bei Angabe von 0 wird individuell lt. stocks_dic eingelesen
#Input - start_year, start_month: wie weit in die Historie zurückgegangen wird (z.b. bis 1995 06)
#Input - end_year, end_month: von welchem Datum die Ermittlung weg erfolgt - wenn year = 0 wird aktuelles Tagesdatum genommen
#Input - sek: Anzahl der Sekunden der Verzögerung bei VPN-Switch
whg = "USD"
index = 0
vpn_land = "no-vpn"
writemodus = 1
index = "s-p_500-index/kursliste"
index = "s-p_500-index/kursliste"
#index = "nasdaq-100-index/kursliste"
#index="dax-30"
#index="tecdax"
sek = 45        #bei 0 Sekunden => kein VPN
start_year = 1989
start_month = 1
end_year = 0
end_month = 0

start_gesamt = timeit.default_timer()
if index != 0:
    stocks_dic = read_index(index)
if end_year == 0:
    end_year = datetime.datetime.now().year
    end_month = datetime.datetime.now().month

vpn_performance = []

abbruch = False
# für jeden Aktientitel aus der Liste Ermittlung einer Zeile mit den Datümern und eine Zeile mit Schlusskursen
start_readstocks = timeit.default_timer()
for stock in stocks_dic:
    if abbruch == True: break

    start_stock = timeit.default_timer()

    # Check ob Aktie bereits im XLS enthalten ist - wenn ja wird nächte Aktie verarbeitet
    if index == 0: check = check_xls(stocks_dic.get(stock), "Stock_Prices.xlsx")
    else: check = check_xls(stocks_dic.get(stock), index.replace ("/", "_").replace ("kursliste", "") + "_Stock_Prices_" + whg + ".xlsx")
    if check == True: continue
    if sek !=0:
        vpn_land = vpn_switch (sek)
        print("Verarbeitung:",stock,"with VPN:",vpn_land,"Aktienkurse lesen...")
    else:
        print ("Verarbeitung:", stock, "without VPN... Aktienkurse lesen...")
    output =  read_prices(stock,start_month, start_year, end_month, end_year, whg)
    if index == 0: save_xls(stocks_dic.get(stock), output, "Stock_Prices.xlsx")
    else: save_xls(stocks_dic.get(stock), output, index.replace ("/", "_").replace ("kursliste", "") + "_Stock_Prices_" + whg + ".xlsx")

    stop_stock = timeit.default_timer ()
    laufzeit = round((stop_stock-start_stock)/60,2)
    print("Laufzeit Aktie ",stock," : ",laufzeit,"min mit Server aus ", vpn_land)
    vpn_performance.append("Laufzeit Aktie "+stock+" : "+str(laufzeit)+"min mit Server aus "+vpn_land)

stop_readstocks = timeit.default_timer()
stop_gesamt = timeit.default_timer()
print("Gesamtlaufzeit: ", round((stop_gesamt-start_gesamt)/60,2), "min")
print("Aktienkurse Gesamt: ", round((stop_readstocks-start_readstocks)/60,2), "min")
print("Statistik VPNs:")
for i in vpn_performance: print(i)
