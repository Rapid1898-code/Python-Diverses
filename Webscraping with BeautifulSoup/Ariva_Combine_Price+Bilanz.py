import xlrd
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side
import pandas as pd

# calculate growth-value for specific value
# start_col => specifies the columns in which the calculation starts
# anzahl_hist => specifies the years for which the mean is build
# row => specifies the row in which the growth should calculated
def calc_growth(start_col,anzahl_hist, row):
    growth_sum = 0
    growth_anz = 0
    if start_col + anzahl_hist < len(row):
        for i in range(0,anzahl_hist):
            if row[start_col+i] not in [""," ","-"] and row[start_col+i+1] not in [""," ","-"]:
                growth_sum = growth_sum + round((row[start_col +i] - row[start_col +i +1]) / row[start_col +i +1] * 100, 2)
            else:
                return False
            growth_anz += 1
        return (round(growth_sum / anzahl_hist, 2))
    else: return False


# read input-excel-sheets
#fn_price = "s-p_500-index__Stock_Prices_USD.xlsx"
#fn_data = "s-p_500-index__Stock_Data.xlsx"
#fn_price = "Stock_PricesNAS.xlsx"
#fn_data = "Stock_DataNAS.xlsx"
fn_price = "Stock_Prices.xlsx"
fn_data = "Stock_Data.xlsx"

wb_price = load_workbook(fn_price)
wb_data = load_workbook(fn_data)

for sh_price in wb_price:
    if sh_price.title == "INDEX": continue
    print("Verarbeitung von"  ,sh_price.title)
    for sh_data in wb_data:
        if sh_price.title.upper() == sh_data.title.upper():  break
    if sh_data["A1"].value == None: continue

    # read price xls in list
    price_list = []
    for row in sh_price.iter_rows():
        zeile = []
        for cell in row:
            if cell.value is None: zeile.append("")
            else: zeile.append(cell.value)
        price_list.append(zeile)

    # read data xls in list
    data_list = []
    for row in sh_data.iter_rows():
        zeile = []
        for cell in row:
            if cell.value is None: zeile.append("")
            else: zeile.append(cell.value)
        data_list.append(zeile)

    # find necessary rows in data-xlsx
    row_title = row_shares = row_netincome_sh = row_revenue_sh = row_totalequity_sh = row_opcashflow_sh = row_dividend_sh = row_totalincome = 0
    for i in range(len(data_list)-1):
        if "Bilanz in Mio." in data_list[i][0]: row_title = data_list[i]
        if "Aktien im Umlauf"  in data_list[i][0]: row_shares= data_list[i]
        if "Ergebnis je Aktie (unverwässert)" == data_list[i][0]: row_netincome_sh = data_list[i]
        if "Umsatz je Aktie" == data_list[i][0]: row_revenue_sh = data_list[i]
        if "Buchwert je Aktie" == data_list[i][0]: row_totalequity_sh = data_list[i]
        if "Cashflow je Aktie" == data_list[i][0]: row_opcashflow_sh = data_list[i]
        if "Dividende je Aktie" == data_list[i][0]: row_dividend_sh = data_list[i]
        if "Gesamtertrag" == data_list[i][0]: row_totalincome = data_list[i]

    # calculate marketcap per day with daily price and yearly outstanding shares
    tmp_year = 0
    idx = 0
    for i in range (1,len(price_list)):
        if price_list[i][0] == "Datum" or price_list[i][0] == "Date": continue
        if datetime.strptime(price_list[i][0],"%d.%m.%Y") != tmp_year:
            for j in range(2,len(row_title)):
                if row_title[j] == "": continue
                if datetime.strptime(price_list[i][0],"%d.%m.%Y").year-1 == int(row_title[j]):
                    tmp_year = int (row_title[j])
                    idx = j
                    break

        price_list[i][1] = round (price_list[i][1], 2)
        if idx == 0: continue
        if price_list[i][1] not in [""," ","-"] and row_shares[idx] not in [""," ","-"]:
            price_list[i][2] = round(price_list[i][1] * row_shares[idx] / 1000,2)
        else:
            price_list[i][2] = "-"
        if len(price_list[i]) <= 3:
            if price_list[i][1] not in [""," ","-"] and row_netincome_sh[idx] not in [""," ","-"]:
                price_list[i][3] = round (price_list[i][1] / row_netincome_sh[idx], 2)
            else: price_list[i][3] = "-"
            if row_revenue_sh != 0:
                if price_list[i][1] not in ["", " ", "-"] and row_revenue_sh[idx] not in ["", " ", "-"]:
                    price_list[i][4] = round (price_list[i][1] / row_revenue_sh[idx], 2)
                else: price_list[i][4] = "-"
            else:
                if price_list[i][1] not in ["", " ", "-"] and row_shares[idx] not in ["", " ", "-"] and row_totalincome[idx] not in ["", " ", "-"]:
                    price_list[i][4] = round (price_list[i][1] * row_shares[idx] / row_totalincome[idx], 2)
                else: price_list[i][4] = "-"
            if price_list[i][1] not in [""," ","-"] and row_totalequity_sh[idx] not in [""," ","-"]:
                price_list[i][5] = round (price_list[i][1] / row_totalequity_sh[idx], 2)
            else: price_list[i][5] = "-"
            if price_list[i][1] not in [""," ","-"] and row_opcashflow_sh[idx] not in [""," ","-"]:
                price_list[i][6] = round (price_list[i][1] / row_opcashflow_sh[idx], 2)
            else: price_list[i][6] = "-"
            if price_list[i][1] not in ["", " ", "-"] and row_dividend_sh[idx] not in ["", " ", "-"]:
                price_list[i][7] = round (row_dividend_sh[idx] / price_list[i][1] * 100, 2)
            else: price_list[i][7] = "-"
            if price_list[i][3] not in [0,""," ","-"]:
                price_list[i][8] = round (1 / price_list[i][3] * 100, 2)
            else: price_list[i][8] = "-"
            tmp_calc = calc_growth (idx, 5, row_netincome_sh)
            if tmp_calc != False:
                price_list[i][9] = round(price_list[i][3] / tmp_calc,2)
            else:
                price_list[i][9] = "-"
        else:
            price_list[i][3] = round (price_list[i][1] / row_netincome_sh[idx], 2)
            if row_revenue_sh != 0:
                if price_list[i][1] not in ["", " ", "-"] and row_revenue_sh[idx] not in ["", " ", "-"]:
                    price_list[i][4] =  round (price_list[i][1] / row_revenue_sh[idx], 2)
                else: price_list[i][4] = "-"
            else:
                if price_list[i][1] not in ["", " ", "-"] and row_shares[idx] not in ["", " ", "-"] and row_totalincome[idx] not in ["", " ", "-"]:
                    price_list[i][4] = round (price_list[i][1] * row_shares[idx] / row_totalincome[idx], 2)
                else: price_list[i][4] = "-"
            if price_list[i][1] not in [""," ","-"] and row_totalequity_sh[idx] not in [""," ","-"]:
                price_list[i][5] = round (price_list[i][1] / row_totalequity_sh[idx], 2)
            else: price_list[i][5] = "-"
            if price_list[i][1] not in [""," ","-"] and row_opcashflow_sh[idx] not in [""," ","-"]:
                price_list[i][6] = round (price_list[i][1] / row_opcashflow_sh[idx], 2)
            else: price_list[i][6] = "-"
            if price_list[i][1] not in ["", " ", "-"] and row_dividend_sh[idx] not in ["", " ", "-"]:
                price_list[i][7] = round (row_dividend_sh[idx] / price_list[i][1] * 100, 2)
            else: price_list[i][7] = "-"
            if price_list[i][3] not in [0,""," ","-"]:
                price_list[i][8] = round (1 / price_list[i][3] * 100, 2)
            else: price_list[i][8] = "-"
            tmp_calc = calc_growth (idx,5,row_netincome_sh)
            if tmp_calc != False:
                price_list[i][9] = round(price_list[i][3] / tmp_calc,2)
            else:
                price_list[i][9] = "-"


    # Überschrift ergänzen
    if price_list[1][0] == "Date" and price_list[2][0] == "Datum": del price_list[1:3]
    price_list.insert(1,["Date","Price","MarketCap in B","PE (price/earnings)","PS (price/sales)", "PB (price/book value",
                         "PC (price/cashflow", "Dividend Yield", "Initial Yield", "PEG Ratio"])
    price_list.insert(2,["Datum","Kurs","MarktKap in Md","KGV (Kurs/Gewinn)", "KUV (Kurs/Umsatz)", "KBV (Kurs/Buchwert)",
                         "KCV (Kurs/Cashflow)", "Dividendenrendite", "Initial Yield", "KGW Kurs/Gewinn/Wachstum"])
    price_list[0][0] = ""

    #  save XLSX
    writer = pd.ExcelWriter (fn_price, engine='openpyxl', options={'strings_to_numbers': True})
    wb_price.remove(sh_price)
    writer.book = wb_price
    pd.DataFrame (price_list).to_excel (writer, sheet_name=sh_price.title.upper(), header=False, index=False)

    # Formatierung XLSX
    column_widths = []
    ws = writer.sheets[sh_price.title.upper()]
    for row in price_list:
        for i, cell in enumerate (row):
            if len (column_widths) > i:
                if len (str (cell)) > column_widths[i]:
                    column_widths[i] = len (str (cell))
            else:
                column_widths += [len (str (cell))]
    for i, column_width in enumerate (column_widths):
        ws.column_dimensions[get_column_letter (i + 1)].width = column_width + 2

    bold = Font (bold=True)
    bg_yellow = PatternFill (fill_type="solid", start_color='fbfce1', end_color='fbfce1')
    bg_grey = PatternFill (fill_type="solid", start_color='babab6', end_color='babab6')
    bg_green = PatternFill (fill_type="solid", start_color='c7ffcd', end_color='fffbc7')
    frame_all = Border (left=Side (style='thin'), right=Side (style='thin'), top=Side (style='thin'),bottom=Side (style='thin'))
    frame_upanddown = Border (top=Side (style='thin'), bottom=Side (style='thin'))

    for cell in ws["A:A"]:
        cell.font = bold
        cell.fill = bg_yellow
        cell.border = frame_all
    for row in ws["1:3"]:
        for cell in row:
            cell.font = bold
            cell.fill = bg_green
            cell.border = frame_all

    freeze = ws["B4"]
    ws.freeze_panes = freeze

while True:
    try:
        writer.save ()
        writer.close ()
        break
    except Exception as e:
        print ("Error: ", e)
        input ("Datei kann nicht geöffnet werden - bitte schließen und <Enter> drücken!")
