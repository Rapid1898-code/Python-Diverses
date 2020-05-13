from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side
import pandas as pd

name_xlsx = "Test.xlsx"
wb =  load_workbook(name_xlsx)
wb._sheets.sort(key=lambda x: x.title)
if "INDEX" in wb.sheetnames: del wb["INDEX"]
wb.create_sheet("INDEX",0)
ws_idx = wb["INDEX"]
link_idx =  name_xlsx + "#" + "INDEX" + "!A1"
ws_idx["A1"] = "INDEX"

column_width = 0
for i, ws in enumerate(wb):
    if ws.title == "INDEX": continue
    if len(ws.title) > column_width: column_width = len(ws.title)
    link = name_xlsx + "#" + ws.title + "!A1"
    ws_idx.cell (row=i+2, column=1).value = '=HYPERLINK("{}", "{}")'.format (link, ws.title)
    wb[ws.title].cell (row=1, column=6).value = '=HYPERLINK("{}", "{}")'.format (link_idx, "Back to INDEX")

#Formatierung
bold = Font (bold=True)
bg_yellow = PatternFill (fill_type="solid", start_color='fbfce1', end_color='fbfce1')
bg_grey = PatternFill (fill_type="solid", start_color='babab6', end_color='babab6')
bg_green = PatternFill (fill_type="solid", start_color='c7ffcd', end_color='fffbc7')
frame_all = Border (left=Side (style='thin'), right=Side (style='thin'), top=Side (style='thin'),
                    bottom=Side (style='thin'))
frame_upanddown = Border (top=Side (style='thin'), bottom=Side (style='thin'))
size14 = Font (bold=True, size="14")

for cell in ws_idx["A:A"]:
    cell.font = bold
    cell.fill = bg_yellow
    cell.border = frame_all
for cell in ws_idx["2:2"]:
    cell.fill = bg_grey
    cell.border = frame_upanddown
for cell in ws_idx["1:1"]:
    cell.font = bold
    cell.fill = bg_green
    cell.border = frame_all
ws_idx["A1"].font = size14
freeze = ws_idx["C2"]
ws_idx.freeze_panes = freeze
ws_idx.column_dimensions["A"].width = column_width

while True:
    try:
        wb.save (name_xlsx)
        wb.close ()
        break
    except Exception as e:
        print ("Error: ", e)
        input ("Datei kann nicht geöffnet werden - bitte schließen und <Enter> drücken!")









