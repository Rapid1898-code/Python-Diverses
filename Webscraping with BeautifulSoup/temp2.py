import requests
import csv
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
import random
import subprocess
import time
import re
import sys
import timeit
from googletrans import Translator

def read_index(index_name, char="00"):
    print("Reading Index",index_name,"starting with Character:",char,"...")
    page_nr=0
    index_stocks = {}
    temp_stocks = {}
    while True:
        page = requests.get ("https://www.ariva.de/"+index_name+"?page="+str(page_nr))
        soup = BeautifulSoup (page.content, "html.parser")
        table  = soup.find(id="result_table_0")
        for row  in table.find_all("td"):
            if row.get("class") == ["ellipsis", "nobr", "new", "padding-right-5"]:
                if row.text.strip().capitalize()[0:2].upper() >= char:
                    index_stocks[row.find("a")["href"][1:]] = row.text.strip().capitalize()
        #Dict sortieren nach Value
        index_stocks = {k: v for k, v in sorted(index_stocks.items(), key=lambda item: item[1])}
        if temp_stocks == index_stocks: break
        page_nr += 1
        temp_stocks = dict(index_stocks)
    print("Finished Reading Index",len(index_stocks), "are read...")
    return(index_stocks)

# index="dax-30"
# index="tecdax"
# index="sdax"
# index="eurostoxx-50"
# index="s-p_500-index/kursliste"
# index="nasdaq-100-index/kursliste"
index="dow-jones-industrial-average"

l=[]
output = read_index(index)
print(output)


