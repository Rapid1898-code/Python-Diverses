from datetime import datetime
from openpyxl import load_workbook
import pandas as pd

fn_price = "dax-30_Stock_Prices_EUR.xlsx"
wb_price = load_workbook(fn_price)
tmp_list = []
for sh_price in wb_price:
    erg_row_date = [sh_price.title]
    erg_row_price = [sh_price.title]
    if sh_price.title in ["INDEX","HowTo"]: continue
    print ("Vearbeitung: ",sh_price.title)
    for row_cells in sh_price.iter_rows(min_col=1, max_col=3):
        if row_cells[0].value not in [None,"Date","Datum"]:
            erg_row_price.insert (1, row_cells[2].value)
            erg_row_date.insert (1, datetime.strptime(row_cells[0].value, "%d.%m.%Y"))
    tmp_list.append(erg_row_date)
    tmp_list.append (erg_row_price)

#Datums-Headline erstellen
erg = []
for i in range(0,len(tmp_list),2):
    for j in range(1,len(tmp_list[i])):
        if tmp_list[i][j] not in erg: erg.append(tmp_list[i][j])
erg.sort()
erg.insert(0,"Datum")

# List der Aktientite mit Datum + Price durchlaufen
for i in range(0, len(tmp_list), 2):
    # List aller Datümer durchlaufen
    print("Finale Verarbeitung: ",tmp_list[i][0])
    for j in range(1, len(erg)-1):
        if erg[j] != tmp_list[i][j]:
            tmp_list[i].insert(j, erg[j])
            tmp_list[i+1].insert(j, "")

end_erg = []
for i in range (1,len(erg)): erg[i] = datetime.strftime (erg[i], "%d.%m.%Y")
end_erg.append(erg)
for i in range(len(tmp_list)):
    if i%2 == 1: end_erg.append(tmp_list[i])

#tmp_list.insert(0,erg)
#for i in range (len(tmp_list)):
#    if i == 0 or i%2 == 1:
#        for j in range(1, len(tmp_list[i])):
#            tmp_list[i][j] = datetime.strftime (tmp_list[i][j], "%d.%m.%Y")

# Schreiben der Liste in ein XLSX
writer = pd.ExcelWriter ("Ariva_MarketCap.xlsx", engine='openpyxl', options={'strings_to_numbers': True})
pd.DataFrame (end_erg).to_excel (writer, sheet_name="MarketCap", header=False, index=False)
while True:
    try:
        writer.save ()
        writer.close ()
        break
    except Exception as e:
        print ("Error: ", e)
        input ("Datei kann nicht geöffnet werden - bitte schließen und <Enter> drücken!")








