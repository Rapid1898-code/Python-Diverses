import requests
import csv
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
import random
import subprocess
import time
import re
import sys
import timeit
from googletrans import Translator


# Ausgabe der Liste als CSV-File inkl. Prüfung ob Datei geöffnet ist
# Input content: ist eine Matrix Liste [[][]]
# Input filenmae: Name des CSVs-File
def csv_write(content, filename):
    while True:
        try:
            with open (filename, "w", newline="") as fp:
                a = csv.writer (fp, delimiter=",")
                a.writerows (content)
                break
        except Exception as e:
            print ("Error: ", e)
            input ("Datei kann nicht geöffnet werden - bitte schließen und <Enter> drücken!")


# Check ob Aktie schon enthalten ist im XLS
def check_xls(stock, filename):
    global writemodus
    try:
        book = load_workbook (filename)
        if writemodus == 0:
            wahl = input (
                "Es befinden sich Daten im Excel-Sheet " + filename + " wollen Sie wirklich die Datei überschreiben (j/n)=")
            if wahl.upper () == "N": writemodus = 1
    except:
        writemodus = 0
        return False
    if stock in book.sheetnames and writemodus == 1:
        print ("Aktie: ", stock, " bereits im XLS: ", filename, " enthalten - Aktie wird übersprungen")
        return True


# Ausgabe der Liste als XLS-File inkl. Prüfung ob Datei geöffnet ist
# Input stock: Name der Aktie
# Input content: Inhalt in Listenform
# Input filenmae: Name des XLSX-File
# Input append: 1=>anhängen von neuen Worksheets, 0=>überschreiben des XLS
def save_xls(stock, content, filename):
    global writemodus
    # check ob append ausgewählt - aber wenn file nicht vorhanden - dann Wechsel über Überschreibmodus 0

    try:
        book = load_workbook (filename)
    except:
        writemodus = 0  # Writemodus auf Neuerstellen 0 setzen, wenn er auf 1 steht aber kein XLS vorhanden ist
    if writemodus == 0:
        writer = pd.ExcelWriter (filename, engine='openpyxl', options={'strings_to_numbers': True})
    elif writemodus == 1:
        book = load_workbook (filename)
        if stock in book.sheetnames:
            print ("Aktie: ", stock, " bereits im XLS: ", filename, " enthalten - Aktie wird übersprungen")
            return
        writer = pd.ExcelWriter (filename, engine='openpyxl', options={'strings_to_numbers': True})
        writer.book = book
    pd.DataFrame (content).to_excel (writer, sheet_name=stock, header=False, index=False)

    # Ausgabe von leerem Arbeitsblatt wenn Inhalt  leer
    if content == []:
        content.append ([])
        while True:
            try:
                writer.save ()
                writer.close ()
                return
            except Exception as e:
                print ("Error: ", e)
                input ("Datei kann nicht geöffnet werden - bitte schließen und <Enter> drücken!")

    # Automatische Anpassung der Spalten nach best fit
    column_widths = []
    ws = writer.sheets[stock]
    # Ermittlung des längsten Wertes pro Spalte
    for row in content:
        for i, cell in enumerate (row):
            if len (column_widths) > i:
                if len (str (cell)) > column_widths[i]:
                    column_widths[i] = len (str (cell))
            else:
                column_widths += [len (str (cell))]
    for i, column_width in enumerate (column_widths):
        # Spalte 2 mit langem Profil fix mit Breite 17 - restliche Spaten immer mit maximalen Wert pro Spalte
        if i == 0:
            ws.column_dimensions[get_column_letter (i + 1)].width = 35
        elif i == 1:
            ws.column_dimensions[get_column_letter (i + 1)].width = 32
        elif 2 <= i <= 10:
            ws.column_dimensions[get_column_letter (i + 1)].width = 15
        else:
            ws.column_dimensions[get_column_letter (i + 1)].width = column_width + 2

    # Formatierung des Excel-Sheets
    bold = Font (bold=True)
    bg_yellow = PatternFill (fill_type="solid", start_color='fbfce1', end_color='fbfce1')
    bg_grey = PatternFill (fill_type="solid", start_color='babab6', end_color='babab6')
    bg_green = PatternFill (fill_type="solid", start_color='c7ffcd', end_color='fffbc7')
    frame_all = Border (left=Side (style='thin'), right=Side (style='thin'), top=Side (style='thin'),
                        bottom=Side (style='thin'))
    frame_upanddown = Border (top=Side (style='thin'), bottom=Side (style='thin'))
    size14 = Font (bold=True, size="14")

    # Formatierung Excel-Sheet
    for i, cont in enumerate (content):
        if cont == []: continue  # Leerzeile überspringen...
        if cont[0].find ("Bilanz in ") != -1:
            row = i + 1
            break
    for cell in ws["A:A"]:
        cell.font = bold
        cell.fill = bg_yellow
        cell.border = frame_all
    for cell in ws["A1:A2"]: cell[0].fill = bg_yellow
    for cell in ws[f"{row}:{row}"]:
        cell.font = bold
        cell.fill = bg_yellow
        cell.border = frame_all
    for cell in ws[f"B{row}:B{len (content)}"]:
        cell[0].fill = bg_yellow
        cell[0].font = bold
        cell[0].border = frame_all
    for cell in ws["C3:C10"]:
        cell[0].fill = bg_yellow
        cell[0].border = frame_all
        cell[0].font = bold
    for cell in ws["D3:D10"]:
        cell[0].fill = bg_yellow
        cell[0].border = frame_all
        cell[0].font = bold
    for cell in ws[f"{row - 1}:{row - 1}"]:
        cell.fill = bg_grey
        cell.border = frame_upanddown
    for cell in ws["2:2"]:
        cell.fill = bg_grey
        cell.border = frame_upanddown
    ws["G3"].font = bold
    ws["G3"].fill = bg_yellow
    ws["G3"].border = frame_all
    ws["H3"].font = bold
    ws["H3"].fill = bg_yellow
    ws["H3"].border = frame_all
    ws["J3"].font = bold
    ws["J3"].fill = bg_yellow
    ws["J3"].border = frame_all
    ws["K3"].font = bold
    ws["K3"].fill = bg_yellow
    ws["K3"].border = frame_all
    for cell in ws["1:1"]:
        cell.font = bold
        cell.fill = bg_green
        cell.border = frame_all
    ws["A1"].font = size14
    freeze = ws["B2"]
    ws.freeze_panes = freeze

    while True:
        try:
            writer.save ()
            writer.close ()
            break
        except Exception as e:
            print ("Error: ", e)
            input ("Datei kann nicht geöffnet werden - bitte schließen und <Enter> drücken!")


# VPN-Switch bei NordVPN mit x Sekunden Verzögerung
# Output: Rückgabe des zufällig gewählten Landes
def vpn_switch(sek):
    countries = ["Austria", "Belgium", "Germany", "Israel", "Italy", "Bosnia and Herzogovina",
                 "Norway", "Poland", "Portugal", "Romania", "Serbia", "Switzerland", "United Kingdom",
                 "Bulgaria", "Croatia", "Estonia", "Finland", "France", "Georgia", "Greece",
                 "Hungary", "Iceland", "Latvia", "Netherlands"]
    rand_country = random.randrange (len (countries) - 1)
    subprocess.call (["C:/Program Files (x86)/NordVPN/NordVPN.exe", "-c", "-g", countries[rand_country]])
    print ("VPN Switch to", countries[rand_country], "...")
    for i in range (sek, 0, -1):
        sys.stdout.write (str (i) + ' ')
        sys.stdout.flush ()
        time.sleep (1)
    print ("Connected to", countries[rand_country], "...")
    return (countries[rand_country])


# Unternehmen eines bestimmten Index werden eingelesen
# Output: Dict in der Form Kürzel von Ariva.de + Name des Titels (z.b '/apple-aktie': 'Apple')
def read_index(index_name, char="00"):
    print("Reading Index",index_name,"starting with Character:",char,"...")
    page_nr=0
    index_stocks = {}
    temp_stocks = {}
    while True:
        page = requests.get ("https://www.ariva.de/"+index_name+"?page="+str(page_nr))
        soup = BeautifulSoup (page.content, "html.parser")
        table  = soup.find(id="result_table_0")
        for row  in table.find_all("td"):
            if row.get("class") == ["ellipsis", "nobr", "new", "padding-right-5"]:
                if row.text.strip().capitalize()[0:2].upper() >= char:
                    index_stocks[row.find("a")["href"][1:]] = row.text.strip().capitalize()
        #Dict sortieren nach Value
        index_stocks = {k: v for k, v in sorted(index_stocks.items(), key=lambda item: item[1])}
        if temp_stocks == index_stocks: break
        page_nr += 1
        temp_stocks = dict(index_stocks)
    print("Finished Reading Index",len(index_stocks), "are read...")
    return(index_stocks)


# calculate growth-value for specific value
# start_col => specifies the columns in which the calculation starts
# anzahl_hist => specifies the years for which the mean is build
# row => specifies the row in which the growth should calculated
def calc_growth(start_col,anzahl_hist, row):
    growth_sum = 0
    growth_anz = 0
    if start_col + anzahl_hist < len(row):
        for i in range(0,anzahl_hist):
            if row[start_col+i] not in [""," ","-"] and row[start_col+i+1] not in [""," ","-"]:
                growth_sum = growth_sum + round((row[start_col +i] - row[start_col +i +1]) / row[start_col +i +1] * 100, 2)
            else:
                return False
            growth_anz += 1
        return (round(growth_sum / anzahl_hist, 2))
    else: return False


# Kennzahlen für das Unternehmen lt. Parameter lesen
# Input Stock: Aktienkennung lt. Ariva.de z.b. /apple-aktie oder /wirecard-aktie
def read_bilanz(stock):
    output_temp = []
    output_gesamt = []
    seite = 0
    jahre_titelleiste = []

    while True:
        link = "https://www.ariva.de/" + stock + "/bilanz-guv?page=" + str (seite) + "#stammdaten"
        try:
            page = requests.get (link)
        except requests.ConnectionError:
            print ("No Connection - Wait für Reconnection...")
            for i in range (60, 0, -1):
                sys.stdout.write (str (i) + ' ')
                sys.stdout.flush ()
                time.sleep (1)
            page = requests.get (link)
        soup = BeautifulSoup (page.content, "html.parser")
        # Kennzahlen auslesen
        table = soup.find_all ("div", class_="column twothirds table")
        for i in table:
            for j in i.find_all ("tr"):
                row = []
                for k in j.find_all ("td"):
                    row.append (k.text.strip ())
                output_temp.append (row)

        # Bereinigung und Formatierung
        for i in range (len (output_temp) - 1, 0, -1):
            # Quartalswerte entfernen
            if "Quartal" in output_temp[i][0]: del output_temp[i]
            # Leerzeilen entfernen + Überschriften Aktive/Passiva
            empty = True
            for j in range (1, len (output_temp[0]) - 1):
                if output_temp[i][j] != "": empty = False
            if empty == True: del output_temp[i]
            # Doppelte Jahreszahlen entfernen
            if i != 0 and output_temp[i] == output_temp[0]: del output_temp[i]

        # Eckdaten Kennzahlen ermitteln und links oben im Sheet speichern
        if seite == 0:
            table = soup.find_all ("h3", class_="arhead undef")
            for i in table:
                if "Bilanz" in i.text and "Geschäftsjahresende" in i.text:
                    output_temp[0][0] = "Bilanz in Mio. " + i.text.strip ()[18:22] + " per " + i.text.strip ()[-7:-1]
                    bilanz_english = "Balance Sheet in M " + i.text.strip ()[18:22] + " per " + i.text.strip ()[-7:-1]

        # Einfügen weiterer Jahreswerte
        if output_temp == [] and output_gesamt == []: return []
        if output_temp[0][1:7] == jahre_titelleiste:
            break
        else:
            jahre_titelleiste = output_temp[0][1:7]
            seite += 6
            if output_gesamt == []:
                output_gesamt = output_temp
            else:
                for i in range (len (output_temp[0]) - 1, 1, -1):
                    if output_temp[0][i] in output_gesamt[0][1:len (output_gesamt[0]) - 1]:
                        continue
                    else:
                        pos_insert = i
                        break
                for i in range (pos_insert, 0, -1):
                    for j in range (len (output_gesamt)):
                        output_gesamt[j].insert (1, output_temp[j][i])
                        if j > 0 and output_gesamt[j][0] != output_temp[j][0]:
                            print (
                                "FEHLER !!! Beschriftung unterschiedlich in Output-Gesamt und Output-Temp in Zeile: ",
                                j)
            output_temp = []

    # Jahresreihenfolge auf absteigend ändern
    for i, cont in enumerate (output_gesamt):
        teil2 = cont[1:]
        teil2.reverse ()
        output_gesamt[i] = cont[0:1] + teil2

    # Leerspalten bereinigen (wo nur "-" enthalten ist
    # Tausender-Punkte entfernen
    for i in range (len (output_gesamt[0]) - 1, 0, -1):
        # Check ob Spalte nur aus Blank oder "-" besteht
        empty = True
        for j in range (len (output_gesamt)):
            if j != 0 and output_gesamt[j][i] not in ["-", ""]: empty = False
            # Tausenderpunkte entfernen
            output_gesamt[j][i] = output_gesamt[j][i].replace (".", "")
            # Umwandlung von Zahlen mit M im Text für Millionen
            if output_gesamt[j][i].find ("M") != -1:
                output_gesamt[j][i] = output_gesamt[j][i].replace (",", ".")
                output_gesamt[j][i] = output_gesamt[j][i].replace ("M", "").strip ()
                output_gesamt[j][i] = round (float (output_gesamt[j][i]) * 1000000, 0)
            # Umwandlung auf float (nur mit "." und nicht mit "," möglich - daher Umwandlung
            try:
                output_gesamt[j][i] = output_gesamt[j][i].replace (",", ".")
                output_gesamt[j][i] = float (output_gesamt[j][i])
            except:
                continue
        if empty == True:
            for j in range (len (output_gesamt)):
                del output_gesamt[j][i]

    # Formate und Texte anpassen
    text_dic = {'Umsatz': 'Revenue', 'Bruttoergebnis vom Umsatz': 'Gross Profit',
                'Operatives Ergebnis (EBIT)': 'EBIT Earning Before Interest & Tax',
                'Finanzergebnis': 'Financial Result',
                'Ergebnis vor Steuer (EBT)': 'EBT Earning Before Tax',
                'Steuern auf Einkommen und Ertrag': 'Taxes on income and earnings',
                'Ergebnis nach Steuer': 'Earnings after tax', 'Minderheitenanteil': 'Minority Share',
                'Jahresüberschuss/-fehlbetrag': 'Net Profit', 'Summe Umlaufvermögen': 'Current Assets',
                'Summe Anlagevermögen': 'Fixed Assets', 'Summe Aktiva': 'Total Assets',
                'Summe kurzfristiges Fremdkapital': 'Short-Term Debt',
                'Summe langfristiges Fremdkapital': 'Long-Term Debt',
                'Summe Fremdkapital': 'Total Liabilities', 'Minderheitenanteil': 'Minority Share',
                'Summe Eigenkapital': 'Equity', 'Summe Passiva': 'Liabilities & Shareholder Equity',
                'Mio.Aktien im Umlauf': 'Million shares outstanding',
                'Gezeichnetes Kapital (in Mio.)': 'Subscribed Capital in M',
                'Ergebnis je Aktie (brutto)': 'Earnings per share',
                'Ergebnis je Aktie (unverwässert)': 'Basic Earnings per share',
                'Ergebnis je Aktie (verwässert)': 'Diluted Earnings per share',
                'Dividende je Aktie': 'Dividend per share',
                'Dividendenausschüttung in Mio': 'Dividend Payment in M',
                'Umsatz je Aktie': 'Revenue per share', 'Buchwert je Aktie': 'Book value per share',
                'Cashflow je Aktie': 'Cashflow per share', 'Bilanzsumme je Aktie': 'Total assets per share',
                'Personal am Ende des Jahres': 'Staff at the end of year',
                'Personalaufwand in Mio. EUR': 'Personnel expenses in M',
                'Personalaufwand in Mio. USD': 'Personnel expenses in M',
                'Aufwand je Mitarbeiter in USD': 'Effort per employee',
                'Umsatz je Mitarbeiter in USD': 'Turnover per employee',
                'Bruttoergebnis je Mitarbeiter in USD': 'Gross Profit per employee',
                'Aufwand je Mitarbeiter in EUR': 'Effort per employee',
                'Umsatz je Mitarbeiter in EUR': 'Turnover per employee',
                'Bruttoergebnis je Mitarbeiter in EUR': 'Gross Profit per employee',
                'KGV (Kurs/Gewinn)': 'PE (price/earnings)',
                'KUV (Kurs/Umsatz)': 'PS (price/sales)', 'KBV (Kurs/Buchwert)': 'PB (price/book value)',
                'KCV (Kurs/Cashflow)': 'PC (price/cashflow)', 'Dividendenrendite in %': 'Dividend Yield in %',
                'Gewinnrendite in %': 'Return on profit in %', 'Eigenkapitalrendite in %': 'Return on Equity in %',
                'Umsatzrendite in %': 'Return on sales in %',
                'Gesamtkapitalrendite in %': 'Total Return on Investment in %',
                'Return on Investment in %': 'Return on Investment in %',
                'Arbeitsintensität in %': 'Work Intensity in %',
                'Eigenkapitalquote in %': 'Equity Ratio in %', 'Fremdkapitalquote in %': 'Debt Ratio in %',
                'Working Capital in Mio': 'Working Capital in M',
                'Gewinn je Mitarbeiter in EUR': 'Earnings per employee',
                'Verschuldungsgrad in %': 'Finance Gearing in %',
                'Gewinn je Mitarbeiter in USD': 'Earnings per employee',
                'Ertrag': 'Income', 'Ertrag je Mitarbeiter in EUR': 'Income per employee',
                'Ertrag je Mitarbeiter in USD': 'Income per employee',
                'Gesamtertrag': 'Total Income', 'Sonderdividende je Aktie': 'Special Dividend per share'}

    # Text bereinigen und englische Spalte ergänzen
    text_dic[output_gesamt[0][0]] = bilanz_english
    umsatz_count = 0
    for i in range (len (output_gesamt)):
        if output_gesamt[i][0] == "Summe Anlagevermögen (*)": output_gesamt[i][0] = "Summe Anlagevermögen"
        if "Working Capital" in output_gesamt[i][0]: output_gesamt[i][0] = "Working Capital in Mio"
        if "Aktien im Umlauf" in output_gesamt[i][0]: output_gesamt[i][0] = "Mio.Aktien im Umlauf"
        if "Umsatz" in output_gesamt[i][0] and umsatz_count == 2:
            output_gesamt[i][0] = "Umsatz je Aktie"
            umsatz_count = 99
        if "Umsatz" in output_gesamt[i][0] and umsatz_count != 2: umsatz_count += 1

        if output_gesamt[i][0] in text_dic:
            output_gesamt[i].insert (1, text_dic.get (output_gesamt[i][0]))
        else:
            output_gesamt[i].insert (1, "")

    # Kennzahlen ergänzen
    revenue_row = grossprofit_row = netincome_row = equity_row = ebit_row = totalassets_row = shorttermdebt_row = 0
    currentassets_row = noncurrentassets_row = currentliabilities_row = cashflow_sh_row = shares_row = 0
    rows = []
    row_add = 29
    for i in range (100): rows.append ([])

    for i_idx, i_cont in enumerate (output_gesamt):
        if i_cont[1] == "Revenue": revenue_row = i_idx
        if i_cont[1] == "Gross Profit": grossprofit_row = i_idx
        if i_cont[1] == "Net Profit": netincome_row = i_idx
        if i_cont[1] == "Equity": equity_row = i_idx
        if i_cont[1] == "EBIT Earning Before Interest & Tax": ebit_row = i_idx
        if i_cont[1] == "Total Assets": totalassets_row = i_idx
        if i_cont[1] == "Short-Term Debt": shorttermdebt_row = i_idx
        if i_cont[1] == "Current Assets": currentassets_row = i_idx
        if i_cont[1] == "Fixed Assets": noncurrentassets_row = i_idx
        if i_cont[1] == "Short-Term Debt": currentliabilities_row = i_idx
        if i_cont[1] == "PC (price/cashflow)": cashflow_sh_row = i_idx
        if i_cont[1] == "Million shares outstanding": shares_row = i_idx
        if i_cont[1] == "PE (price/earnings)": earnings_sh_row = i_idx
        if i_cont[1] == "EBIT Earning Before Interest & Tax": ebit_row = i_idx

    if revenue_row != 0 and grossprofit_row != 0: rows[0] = ["Bruttoergebnis Marge in %", "Gross Profit Marge in %"]
    if currentassets_row != 0 and totalassets_row != 0: rows[1] = ["Kurzfristige Vermögensquote in %", "Current Assets Ratio in %"]
    if netincome_row != 0 and revenue_row != 0: rows[2] = ["Nettogewinn Marge in %", "Net Profit Marge in %"]
    if ebit_row != 0 and revenue_row != 0: rows[3] = ["Operative Ergebnis Marge in %", "EBIT Marge in %"]
    if revenue_row != 0 and totalassets_row != 0: rows[4] = ["Vermögensumsschlag in %", "Asset Turnover in %"]
    if noncurrentassets_row != 0 and totalassets_row != 0: rows[5] = ["Langfristige Vermögensquote in %","Non-Current Assets Ratio in %"]
    if netincome_row != 0 and totalassets_row != 0: rows[6] = ["Gesamtkapitalrentabilität", "ROA Return on Assets in %"]
    if ebit_row != 0 and totalassets_row != 0: rows[7] = ["Ertrag des eingesetzten Kapitals","ROCE Return on Cap. Empl. in %"]
    if equity_row != 0 and noncurrentassets_row != 0: rows[8] = ["Eigenkapital zu Anlagevermögen","Equity to Fixed Assets in %"]
    if currentassets_row != 0 and currentliabilities_row != 0: rows[9] = ["Liquidität Dritten Grades","Current Ratio in %"]
    if cashflow_sh_row != 0 and shares_row != 0: rows[10] = ["Operativer Cashflow", "Operating Cashflow in M"]
    if shares_row != 0: rows[11] = ["Aktienrückkauf", "Share Buyback in M"]
    if revenue_row != 0 :
        rows[12] = ["Umsatzwachstum 1J in %", "Revenue Growth 1Y in %"]
        rows[13] = ["Umsatzwachstum 3J in %", "Revenue Growth 3Y in %"]
        rows[14] = ["Umsatzwachstum 5J in %", "Revenue Growth 5Y in %"]
        rows[15] = ["Umsatzwachstum 10J in %", "Revenue Growth 10Y in %"]
    if netincome_row != 0 :
        rows[16] = ["Gewinnwachstum 1J in %", "Earnings Growth 1Y in %"]
        rows[17] = ["Gewinnwachstum 3J in %", "Earnings Growth 3Y in %"]
        rows[18] = ["Gewinnwachstum 5J in %", "Earnings Growth 5Y in %"]
        rows[19] = ["Gewinnwachstum 10J in %", "Earnings Growth 10Y in %"]
    if earnings_sh_row !=0 and netincome_row != 0: rows[20] = ["PEG Ratio", "KGW Kurs/Gewinn/Wachstum"]
    if ebit_row != 0 :
        rows[21] = ["EBIT-Wachstum 1J in %", "EBIT Growth 1Y in %"]
        rows[22] = ["EBIT-Wachstum 3J in %", "EBIT Growth 3Y in %"]
        rows[23] = ["EBIT-Wachstum 5J in %", "EBIT Growth 5Y in %"]
        rows[24] = ["EBIT-Wachstum 10J in %", "EBIT Growth 10Y in %"]
    if cashflow_sh_row != 0:
        rows[25] = ["Op.Cashflow Wachstum 1J in %", "Op.Cashflow Wachstum 1Y in %"]
        rows[26] = ["Op.Cashflow Wachstum 3J in %", "Op.Cashflow Wachstum 3Y in %"]
        rows[27] = ["Op.Cashflow Wachstum 5J in %", "Op.Cashflow Wachstum 5Y in %"]
        rows[28] = ["Op.Cashflow Wachstum 10J in %", "Op.Cashflow Wachstum 10Y in %"]

    for i in range (2, len (output_gesamt[len (output_gesamt) - 1]) - 1):
        for j in  range (row_add):
            if j == 0 and rows[j] != []:
                if output_gesamt[grossprofit_row][i] not in [""," ","-"] and output_gesamt[revenue_row][i] not in [""," ","-"]:
                    rows[j].append (round (output_gesamt[grossprofit_row][i] / output_gesamt[revenue_row][i] * 100, 2))
                else: rows[j].append("-")
            if j == 1 and rows[j] != []:
                if output_gesamt[currentassets_row][i] not in [""," ","-"] and output_gesamt[totalassets_row][i] not in [""," ","-"]:
                    rows[j].append (round (output_gesamt[currentassets_row][i] / output_gesamt[totalassets_row][i] * 100, 2))
                else: rows[j].append("-")
            if j == 2 and rows[j] != []:
                if output_gesamt[netincome_row][i] not in [""," ","-"] and output_gesamt[revenue_row][i] not in [""," ","-"]:
                    rows[j].append (round (output_gesamt[netincome_row][i] / output_gesamt[revenue_row][i] * 100, 2))
                else: rows[j].append("-")
            if j == 3 and rows[j] != []:
                if output_gesamt[ebit_row][i] not in [""," ","-"] and output_gesamt[revenue_row][i] not in [""," ","-"]:
                    rows[j].append (round (output_gesamt[ebit_row][i] / output_gesamt[revenue_row][i] * 100, 2))
                else: rows[j].append ("-")
            if j == 4 and rows[j] != []:
                if output_gesamt[revenue_row][i] not in [""," ","-"] and output_gesamt[totalassets_row][i] not in [""," ","-"]:
                    rows[j].append (round (output_gesamt[revenue_row][i] / output_gesamt[totalassets_row][i] * 100, 2))
                else: rows[j].append ("-")
            if j == 5 and rows[j] != []:
                if output_gesamt[noncurrentassets_row][i] not in [""," ","-"] and output_gesamt[totalassets_row][i] not in [""," ","-"]:
                    rows[j].append (round (output_gesamt[noncurrentassets_row][i] / output_gesamt[totalassets_row][i] * 100, 2))
                else: rows[j].append ("-")
            if j == 6 and rows[j] != []:
                if output_gesamt[netincome_row][i] not in [""," ","-"] and output_gesamt[totalassets_row][i] not in [""," ","-"]:
                    rows[j].append (round (output_gesamt[netincome_row][i] / output_gesamt[totalassets_row][i] * 100, 2))
                else: rows[j].append ("-")
            if j == 7 and rows[j] != []:
                if output_gesamt[ebit_row][i] not in [""," ","-"] and output_gesamt[totalassets_row][i] not in [""," ","-"] and output_gesamt[shorttermdebt_row][i]:
                        rows[j].append (round (output_gesamt[ebit_row][i] / (output_gesamt[totalassets_row][i] - output_gesamt[shorttermdebt_row][i]) * 100, 2))
                else: rows[j].append ("-")
            if j == 8 and rows[j] != []:
                if output_gesamt[equity_row][i] not in [""," ","-"] and output_gesamt[noncurrentassets_row][i] not in [""," ","-"]:
                        rows[j].append (round (output_gesamt[equity_row][i] / output_gesamt[noncurrentassets_row][i] * 100, 2))
                else: rows[j].append ("-")
            if j == 9 and rows[j] != []:
                if output_gesamt[currentassets_row][i] not in [""," ","-"] and output_gesamt[currentliabilities_row][i] not in [""," ","-"]:
                    rows[j].append (round (output_gesamt[currentassets_row][i] / output_gesamt[currentliabilities_row][i] * 100, 2))
                else: rows[j].append ("-")
            if j == 10 and rows[j] != []:
                if output_gesamt[cashflow_sh_row][i] not in [""," ","-"] and output_gesamt[shares_row][i] not in [""," ","-"]:
                    rows[j].append (output_gesamt[cashflow_sh_row][i] * output_gesamt[shares_row][i])
                else: rows[j].append ("-")
            if j == 11 and rows[j] != []:
                if i < len(output_gesamt[30]) and isinstance(output_gesamt[shares_row][i],float) and isinstance (output_gesamt[shares_row][i + 1], float):
                    rows[j].append (output_gesamt[shares_row][i+1] - output_gesamt[shares_row][i])
                else:
                    rows[j].append("-")
            if j == 12 and rows[j] != []:
                for k in [1,3,5,10]:
                    tmp_calc = calc_growth (i,k,output_gesamt[revenue_row])
                    if k == 1:
                        if tmp_calc != False: rows[j].append (tmp_calc)
                        else: rows[j].append ("-")
                    if k == 3:
                        if tmp_calc != False: rows[j+1].append (tmp_calc)
                        else: rows[j+1].append ("-")
                    if k == 5:
                        if tmp_calc != False: rows[j+2].append (tmp_calc)
                        else: rows[j+2].append ("-")
                    if k == 10:
                        if tmp_calc != False: rows[j+3].append (tmp_calc)
                        else: rows[j+3].append ("-")
            if j == 16 and rows[j] != []:
                for k in [1,3,5,10]:
                    tmp_calc = calc_growth (i,k,output_gesamt[netincome_row])
                    if k == 1:
                        if tmp_calc != False: rows[j].append (tmp_calc)
                        else: rows[j].append ("-")
                    if k == 3:
                        if tmp_calc != False: rows[j+1].append (tmp_calc)
                        else: rows[j+1].append ("-")
                    if k == 5:
                        if tmp_calc != False:
                            rows[j+2].append (tmp_calc)
                            if output_gesamt[earnings_sh_row][i] not in ["","-"," "]:
                                rows[j+4].append (round(output_gesamt[earnings_sh_row][i] / tmp_calc,2))
                            else:
                                rows[j + 4].append ("-")
                        else:
                            rows[j+2].append ("-")
                            rows[j+4].append ("-")
                    if k == 10:
                        if tmp_calc != False: rows[j+3].append (tmp_calc)
                        else: rows[j+3].append ("-")
            if j == 21 and rows[j] != []:
                for k in [1,3,5,10]:
                    tmp_calc = calc_growth (i,k,output_gesamt[ebit_row])
                    if k == 1:
                        if tmp_calc != False: rows[j].append (tmp_calc)
                        else: rows[j].append ("-")
                    if k == 3:
                        if tmp_calc != False: rows[j+1].append (tmp_calc)
                        else: rows[j+1].append ("-")
                    if k == 5:
                        if tmp_calc != False: rows[j+2].append (tmp_calc)
                        else: rows[j+2].append ("-")
                    if k == 10:
                        if tmp_calc != False: rows[j+3].append (tmp_calc)
                        else: rows[j+3].append ("-")
            if j == 25 and rows[j] != []:
                for k in [1,3,5,10]:
                    tmp_calc = calc_growth (i,k,output_gesamt[cashflow_sh_row])
                    if k == 1:
                        if tmp_calc != False: rows[j].append (tmp_calc)
                        else: rows[j].append ("-")
                    if k == 3:
                        if tmp_calc != False: rows[j+1].append (tmp_calc)
                        else: rows[j+1].append ("-")
                    if k == 5:
                        if tmp_calc != False: rows[j+2].append (tmp_calc)
                        else: rows[j+2].append ("-")
                    if k == 10:
                        if tmp_calc != False: rows[j+3].append (tmp_calc)
                        else: rows[j+3].append ("-")
            if j in [13,14,15,17,18,19,20,22,23,24,26,27,28]: continue

    for i in range (row_add):
        output_gesamt.insert(len(output_gesamt)-1, rows[i])

    return output_gesamt


# Stammdaten für das Unternehmen lt. Parameter lesen
# Input Stock: Aktienkennung lt. Ariva.de z.b. /apple-aktie oder /wirecard-aktie
def read_stamm(stock):
    output = []
    translator = Translator()

    # Stammdaten auslesen
    link = "https://www.ariva.de/" + stock + "/bilanz-guv?page=" + "0" + "#stammdaten"
    try:
        page = requests.get (link)
    except requests.ConnectionError:
        print ("No Connection - Wait für Reconnection...")
        for i in range (60, 0, -1):
            sys.stdout.write (str (i) + ' ')
            sys.stdout.flush ()
            time.sleep (1)
        page = requests.get (link)
    soup = BeautifulSoup (page.content, "html.parser")

    output.append (["STAMMDATEN / BASE DATA", ""])
    table = soup.find_all ("div", class_="column half")
    for i in table:
        for j in i.find_all ("tr"):
            row = []
            break_jn = False
            translate_jn = False
            for k in j.find_all ("td"):
                if k.text.strip () == "":
                    row.append ("-")
                elif k.text.strip () == "Gründungsjahr":
                    row.append ("Gründungsjahr / Founding Year")
                elif k.text.strip () == "Gelistet seit":
                    row.append ("Gelistet Seit / Listed Since")
                elif k.text.strip () == "Nennwert / Aktie":
                    row.append ("Nominalwert / Nominal Value")
                elif k.text.strip () == "Land":
                    row.append ("Land / Country")
                elif k.text.strip () == "Währung":
                    row.append ("Währung / Currency")
                elif k.text.strip () == "Branche":
                    row.append ("Branche / Industry")
                    translate_jn = True
                elif k.text.strip () == "Aktientyp":
                    break_jn = True
                    break
                    #row.append ("Aktientyp / Share Type")
                elif k.text.strip () == "Sektor":
                    row.append ("Sektor / Sector")
                    translate_jn = True
                elif k.text.strip () == "Gattung":
                    row.append ("Typ / Genre")
                elif k.text.strip () == "Adresse":
                    row.append ("Adresse / Address")
                elif k.text.strip () == "Profil":
                    row.append ("Beschreibung / Profile")
                else:
                    if translate_jn == True:
                        row.append(translator.translate (k.text.strip (), src="de", dest="en").text.title())
                        translate_jn = False
                    else:
                        row.append(k.text.strip ())
            if break_jn == False: output.append (row)

    # Kontakte auslesen
    output[0].extend (["KONTAKT / CONTACT", "", ""])
    table = soup.find_all ("div", class_="column half last")
    nr = 1
    for i in table:
        for j in i.find_all ("tr"):
            # Adresse wird unten vor der Zeile Management eingefügt, weil sie zu groß ist
            next_one_adress = False
            for k in j.find_all ("td"):
                if next_one_adress == True:
                    next_one_adress = False
                    if k.text.strip () == "":
                        next_one_adress_row.append ("-")
                    else:
                        next_one_adress_row.append (k.text.strip ())
                    continue
                if k.text.strip () == "Adresse":
                    next_one_adress = True
                    next_one_adress_row = ["Adresse / Address"]
                    nr -= 1
                    continue
                if k.text.strip () == "":
                    output[nr].append ("-")
                elif k.text.strip () == "Telefon":
                    output[nr].extend (["Telefon / Phone", ""])
                elif k.text.strip () == "Fax":
                    output[nr].extend (["Fax", ""])
                elif k.text.strip () == "Internet":
                    output[nr].extend (["Internet", ""])
                elif k.text.strip () == "E-Mail":
                    output[nr].extend (["E-Mail", ""])
                elif k.text.strip () == "IR Telefon":
                    output[nr].extend (["Inv. Relations Telefon / Phone", ""])
                elif k.text.strip () == "IR E-Mail":
                    output[nr].extend (["Inv. Relations E-Mail", ""])
                elif k.text.strip () == "Kontaktperson":
                    output[nr].extend (["Kontaktperson / Contact Person", ""])
                else:
                    output[nr].append (k.text.strip ())
            output[nr].append ("")
            nr += 1

    # Termine auslesen
    table = soup.find_all ("div", class_="termine abstand new")
    nr = 1
    # Termine - Überschrift auslesen
    for i in table:
        cont = i.find ("h3", class_="arhead undef").text.strip ().upper ()
        output[0].extend ([cont.replace ("TERMINE","TERMINE / EVENTS"), "", ""])
    # Termine Inhalt auslesen
    for i in table:
        for j in i.find_all ("tr"):
            for k in j.find_all ("td"):
                # Prüfung ob Tag.Monat damit Jahr ergänzt wird
                pattern = '^[0-9][0-9].[0-9][0-9].$'
                if re.match (pattern, k.text.strip ()):
                    output[nr].append (k.text.strip () + cont[-4:])
                else:
                    output[nr].append (translator.translate (k.text.strip(), src="de", dest="en").text.title())
            output[nr].append ("")
            nr += 1

    # Aktionäre
    output[0].extend (["AKTIONÄRE / SHAREHOLDERS", ""])
    table = soup.find_all ("div", class_="aktStruktur abstand new")
    nr = 1
    for i in table:
        for j in i.find_all ("tr"):
            for k in j.find_all ("td"):
                if nr <= 10:
                    if k.text.strip ().find ("%") != -1:
                        while (len (output[nr]) < 11): output[nr].append ("")
                        output[nr].append (k.text.strip ())
                    else:
                        while (len (output[nr]) < 9): output[nr].append ("")
                        output[nr].append (k.text.strip ())
            nr += 1

    # Adresse von oben einfügen
    output.append (next_one_adress_row)

    # Management / Aufsichtsrat
    table = soup.find_all ("div", class_="management abstand new")
    for i in table:
        for j in i.find_all ("tr"):
            row = []
            for k in j.find_all ("td"):
                if k.text.strip () == "Aufsichtsrat": row.append ("Aufsichtsrat / Board")
                else: row.append (k.text.strip ())
            output.append (row)

    # Profil
    table = soup.find (id="profil_text")
    txt = ""
    for i in table.find_all ("p"):
        if txt == "":
            txt = txt + i.text.strip ()
        else:
            txt = txt + " " + i.text.strip ()
    output.append (["Beschreibung", txt])
    output.append (["Profile", translator.translate (txt, src="de", dest="en").text])

    # Aktien Kennzeichnungen lesen und als Titel einfügen
    link = "https://www.ariva.de/" + stock
    try:
        page = requests.get (link)
    except requests.ConnectionError:
        print ("No Connection - Wait für Reconnection...")
        for i in range (60, 0, -1):
            sys.stdout.write (str (i) + ' ')
            sys.stdout.flush ()
            time.sleep (1)
        page = requests.get (link)
    soup = BeautifulSoup (page.content, "html.parser")
    table = soup.find_all ("div", class_="snapshotHeader abstand")
    for i in table:
        for j in i.find_all ("div", class_="verlauf snapshotInfo"):
            tmp_str = j.text.strip ().replace ("\n", "").replace ("\t", "")
    output.insert (0, [stock.upper ().replace ("-", " ").replace ("AKTIE", "").replace ("_", " "), tmp_str])
    output.insert (1, [""])

    return (output)


#stocks_dic = {'apple-aktie': 'Apple'}
stocks_dic = {'apple-aktie': 'Apple', 'infineon-aktie': 'Infineon'}
# stocks_dic = {'continental-aktie': 'Continental', 'apple-aktie': 'Apple'}
# stocks_dic = {'zimmer_holdings-aktie': 'Zimmer Holdings', 'zebra_technologies-aktie': 'Zebra Technologies'}
# stocks_dic = {'alaska_air_group-aktie': 'Alaska','alliance_data_systems-aktie': 'Alliance'}

# Input-Parameter
# Input - Angabe welcher Index gelesen werden soll (z.B. DAX-30) - bei Angabe von 0 wird individuell lt. stocks_dic eingelesen
# Input - sek: Anzahl der Sekunden der Verzögerung bei VPN-Switch
index = 0
char_index = "FA"
vpn_land = "no-vpn"
writemodus = 1
# index="dax-30"
# index="tecdax"
# index="sdax"
# index="eurostoxx-50"
index="s-p_500-index/kursliste"
# index="nasdaq-100-index/kursliste"

sek = 45        #bei 0 Sekunden => kein VPN
entry = 30      #Wechsel der VPN-Verbindung bei allen 20 Einträgen

start_readstocks = timeit.default_timer ()
if index != 0:
    stocks_dic = read_index(index,char_index.upper())
for i, stock in enumerate(stocks_dic):
    if sek != 0:
        if i%30 == 0: vpn_land = vpn_switch (sek)
        print ("Verarbeitung:", stock, "with VPN:", vpn_land)
    else:
        print ("Verarbeitung:", stock, "without VPN... Aktienkurse lesen...")
    if index == 0:
        check = check_xls (stocks_dic.get (stock), "Stock_Data.xlsx")
    else:
        check = check_xls (stocks_dic.get (stock),
                           index.replace ("/", "_").replace ("kursliste", "") + "_Stock_Data.xlsx")
    if check == True: continue
    output = read_bilanz (stock)
    if output != []:
        output_stamm = read_stamm (stock)
        output.insert (0, [])
        for i in range (len (output_stamm) - 1, -1, -1):
            output.insert (0, output_stamm[i])
    if index == 0:
        save_xls (stocks_dic.get (stock), output, "Stock_Data.xlsx")
    else:
        save_xls (stocks_dic.get (stock), output,
                  index.replace ("/", "_").replace ("kursliste", "") + "_Stock_Data.xlsx")
    if writemodus == 0: writemodus = 1

# Worksheets sortieren im XLSX
if index == 0: name_xlsx = "Stock_Data.xlsx"
else: name_xlsx = index.replace ("/", "_").replace ("kursliste", "") + "_Stock_Data.xlsx"
wb =  load_workbook(name_xlsx)
wb._sheets.sort(key=lambda x: x.title)
while True:
    try:
        wb.save (name_xlsx)
        wb.close ()
        break
    except Exception as e:
        print ("Error: ", e)
        input ("Datei kann nicht geöffnet werden - bitte schließen und <Enter> drücken!")

stop_readstocks = timeit.default_timer ()
print ("Verarbeitung beendet - Gesamtlaufzeit: ", round ((stop_readstocks - start_readstocks) / 60, 2), "min")
