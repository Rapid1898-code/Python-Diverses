from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import pandas as pd

wb = load_workbook("Test.xlsx")
writer = pd.ExcelWriter ("Test.xlsx", engine='openpyxl', options={'strings_to_numbers': True})
writer.book = wb
ws = wb["TAB"]

bg_yellow = PatternFill (fill_type="solid", start_color='fbfce1', end_color='fbfce1')

for row in ws["A1":"B5"]:
    for cell in row:
        cell.fill = bg_yellow

while True:
    try:
        writer.save ()
        writer.close ()
        break
    except Exception as e:
        print ("Error: ", e)
        input ("Datei kann nicht geöffnet werden - bitte schließen und <Enter> drücken!")
