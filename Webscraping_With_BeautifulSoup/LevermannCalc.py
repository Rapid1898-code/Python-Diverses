import YahooCrawler
from datetime import datetime, timedelta
from datetime import date
import timeit
import calendar
import xlwings as xw
import logging
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import pandas as pd

def save_xls(stock, content, filename):
    book = load_workbook (filename)
    writer = pd.ExcelWriter (filename, engine='openpyxl', options={'strings_to_numbers': True})
    writer.book = book
    pd.DataFrame (content).to_excel (writer, sheet_name=stock, header=False, index=False)

    # Automatische Anpassung der Spalten nach best fit
    column_widths = []
    ws = writer.sheets[stock]
    # Ermittlung des längsten Wertes pro Spalte
    for row in content:
        for i, cell in enumerate (row):
            if len (column_widths) > i:
                if len (str (cell)) > column_widths[i]:
                    column_widths[i] = len (str (cell))
            else:
                column_widths += [len (str (cell))]
    for i, column_width in enumerate (column_widths):
        # Spalte 2 mit langem Profil fix mit Breite 17 - restliche Spaten immer mit maximalen Wert pro Spalte
        ws.column_dimensions[get_column_letter (i + 1)].width = column_width + 2

    # Formatierung des Excel-Sheets
    bold = Font (bold=True)
    bg_yellow = PatternFill (fill_type="solid", start_color='fbfce1', end_color='fbfce1')
    bg_grey = PatternFill (fill_type="solid", start_color='babab6', end_color='babab6')
    bg_green = PatternFill (fill_type="solid", start_color='8af542', end_color='8af542')
    bg_blue = PatternFill (fill_type="solid", start_color='42b9f5', end_color='42b9f5')
    bg_orange = PatternFill (fill_type="solid", start_color='fc6f03', end_color='fc6f03')

    frame_all = Border (left=Side (style='thin'), right=Side (style='thin'), top=Side (style='thin'),
                        bottom=Side (style='thin'))
    frame_upanddown = Border (top=Side (style='thin'), bottom=Side (style='thin'))
    size14 = Font (bold=True, size="14")
    size12 = Font (bold=True, size="12")
    left_allign = Alignment (horizontal="left")
    right_allign = Alignment (horizontal="right")

    for row in ws["D1":"G34"]:
        for cell in row: cell.alignment = right_allign
    areas = ["A7:G19","A27:G31"]
    for area in areas:
        for row in ws[area]:
            for cell in row: cell.border = frame_all
    for row in ws["A1":"B4"]:
        for cell in row:
            cell.fill = bg_yellow
            cell.font = size12
            cell.alignment = left_allign
    areas = ["A7:C19","A27:C31"]
    for area in areas:
        for row in ws[area]:
            for cell in row:
                cell.fill = bg_green
                cell.font = size12
                cell.alignment = left_allign
    for i in ["A6","D6","E6","A26","D26","E26"]:
        ws[i].fill = bg_green
        ws[i].font = size12
    for i in ["G20","G32"]:
        ws[i].fill = bg_blue
        ws[i].font = size14
    areas = ["B6:C6","B26:C26"]
    for area in areas:
        for row in ws[area]:
            for cell in row:
                cell.fill = bg_blue
                cell.font = size12
    if "Buy" in rec:
        for row in ws["B20":"C22"]:
            for cell in row:
                cell.fill = bg_blue
                cell.font = size14
    if "Hold" in rec:
        for row in ws["B20":"C22"]:
            for cell in row:
                cell.fill = bg_yellow
                cell.font = size14
    if "Sell" in rec:
        for row in ws["B20":"C22"]:
            for cell in row:
                cell.fill = bg_orange
                cell.font = size14
    if "Buy" in rec_light:
        for row in ws["B32":"C34"]:
            for cell in row:
                cell.fill = bg_blue
                cell.font = size14
    if "Sell" in rec_light:
        for row in ws["B32":"C34"]:
            for cell in row:
                cell.fill = bg_orange
                cell.font = size14

    while True:
        try:
            writer.save ()
            writer.close ()
            break
        except Exception as e:
            print ("Error: ", e)
            input ("Datei kann nicht geöffnet werden - bitte schließen und <Enter> drücken!")

if __name__ == '__main__':
    logging.basicConfig(filename="LevermannCalc.log", format='%(asctime)s %(message)s', filemode='w')     #Create and configure logger
    logger = logging.getLogger ()       # Creating an object
    logger.setLevel (logging.DEBUG)     # Setting the threshold of logger to DEBUG
    inp = "LevermannInput.xlsx"
    out = "LevermannScores.xlsx"

    wb = xw.Book(inp)
    ws_db = wb.sheets["Input"]
    path = os.getcwd()
    fn = path + "\\" + inp
    logger.info("Filename:" + fn)
    l_stocks = ws_db.range("A3:A100").value
    l_index = ws_db.range("B3:B100").value
    l_fin_flag = ws_db.range("C3:C100").value
    l_index_selection = ws_db.range("E2:E100").value

    for idx_stock, stock in enumerate(l_stocks):
        if stock == None: break

        cell_list = ws_db.range ("A" + str(idx_stock+3) + ":" + "C" + str(idx_stock+3)).value
        index = cell_list[1]
        fin = cell_list[2]

        if index == None or fin == None:
            print("Stock:",stock,"skipped - no index or finance flag choosen...")
            continue
        if fin.upper() not in ["Y","J","N"]:
            print("Stock:",stock,"skipped - wrong finance flag in cell C",idx_stock+3,"...")
            continue
        if index not in l_index_selection:
            print("Stock:",stock,"skipped - wrong index name in cell B",idx_stock+3,"...")
            continue

        print("Working on...")

        start = timeit.default_timer()

        #1 - Return On Equity RoE / Eigenkapitalrendite
        stat1,stat2 = YahooCrawler.read_yahoo_statistics(stock)
        if stat1["Return on Equity (ttm)"] != "N/A":
            roe = float(stat1["Return on Equity (ttm)"].replace("%",""))
        else: roe = "N/A"
        marketcap = stat2["Market Cap (intraday)"]
        shares_outstanding = stat1["Shares Outstanding"]

        hist_price_stock = YahooCrawler.read_yahoo_histprice(stock)
        hist_price_index = YahooCrawler.read_yahoo_histprice(index)

        #2 - EBIT-Margin / EBIT Marge
        insstat = YahooCrawler.read_yahoo_income_statement (stock)
        if fin.upper() == "N":
            ebit = insstat["EBIT"][0]
            revenue = insstat["Total Revenue"][0]
            ebit_marge = round(ebit / revenue * 100,2)
        else: ebit_marge = "N/A"

        #4 - P/E-Ratio History 5Y / KGV Historisch 5J
        net_income = insstat["Net Income"]
        count = eps_hist = 0
        pe_ratio_hist_list = []
        for idx,cont in enumerate(net_income):
            if cont == "-": continue
            elif insstat["Breakdown"][idx] == "ttm": continue
            else:
                dt1 = datetime.strptime(insstat["Breakdown"][idx],"%m/%d/%Y")
                dt2 = datetime.strftime(dt1, "%Y-%m-%d")
                tmp_date, tmp_price = YahooCrawler.read_dayprice(hist_price_stock,dt2,"+")
                #DEBUG-INFO
                #print("Price: ", tmp_price)
                #print("NetIncome: ", cont)
                #print("Shares Outstanding: ", shares_outstanding)
                #print("\n")

                eps_hist += tmp_price / (cont / shares_outstanding)
                pe_ratio_hist_list.append(round(tmp_price / (cont / shares_outstanding),2))
                count += 1
        #DEBUG-INFO
        #print("EPS-Hist-Summe: ", eps_hist)
        #print("EPS-Hist-Count: ", count)
        #print(pe_ratio_hist)
        pe_ratio_hist = round(eps_hist / count,2)

        #3 - Equity Ratio / Eigenkaptialquote
        bal_sheet = YahooCrawler.read_yahoo_balance_sheet(stock)
        equity = bal_sheet["Stockholders' Equity"][0]
        total_assets = bal_sheet["Total Assets"][0]
        eq_ratio = round(equity / total_assets * 100,2)

        #5 - P/E-Ratio Actual / KGV Aktuell
        summary = YahooCrawler.read_yahoo_summary(stock)
        pe_ratio = float(summary["pe_ratio"])
        name = summary["name"]

        #6 - Analyst Opinions / Analystenmeinung
        analyst_rating = YahooCrawler.read_zacks_rating(stock)
        if analyst_rating["Rating"] == "N/A": rating = "N/A"
        else: rating = analyst_rating["Rating"][0]

        #7 Reaction to quarter numbers / Reaktion auf Quartalszahlen
        dates_earnings = YahooCrawler.read_yahoo_earnings_cal(stock)
        #DEBUG INFO
        #print(dates_earnings)
        for key in sorted(dates_earnings.keys(), reverse=True):
            if datetime.strptime(key,"%Y-%m-%d") < datetime.today(): break
        last_earningsinfo = key
        stock_price_before = YahooCrawler.read_dayprice(hist_price_stock,key,"+")
        dt1 = datetime.strptime(stock_price_before[0],"%Y-%m-%d") + timedelta (days=1)
        dt2 = datetime.strftime (dt1, "%Y-%m-%d")
        stock_price_after = YahooCrawler.read_dayprice(hist_price_stock,dt2,"+")
        index_price_before = YahooCrawler.read_dayprice(hist_price_index,stock_price_before[0],"+")
        index_price_after = YahooCrawler.read_dayprice(hist_price_index,dt2,"+")
        stock_reaction = ((stock_price_after[1]-stock_price_before[1])/stock_price_before[1])*100
        index_reaction = ((index_price_after[1]-index_price_before[1])/index_price_before[1])*100
        reaction = round(stock_reaction - index_reaction,2)

        #8 Profit Revision / Gewinnrevision
        #13 - Profit Growth / Gewinnwachstum
        analysis = YahooCrawler.read_yahoo_analysis(stock)
        next_year_est_current = float(analysis["Current Estimate"][3])
        next_year_est_90d_ago = float(analysis["90 Days Ago"][3])
        profit_revision = next_year_est_current - next_year_est_90d_ago
        if next_year_est_90d_ago == 0:
            profit_revision = 0
        else:
            profit_revision = round(((next_year_est_current-next_year_est_90d_ago)/next_year_est_90d_ago)*100,2)
        profit_growth_act = float(analysis["Current Estimate"][2])
        profit_growth_fut = float(analysis["Current Estimate"][3])
        if profit_growth_act == 0:
            profit_growth = 0
        else:
            profit_growth = round(((profit_growth_fut - profit_growth_act) / profit_growth_act)*100,2)

        #9 Price Change 6month / Kurs Heute vs. Kurs vor 6M
        #10 Price Change 12month / Kurs Heute vs. Kurs vo 1J
        #11 Price Momentum / Kursmomentum Steigend
        dt1 = datetime.strftime (datetime.today(), "%Y-%m-%d")
        dt2 = datetime.today() - timedelta (days=180)
        dt2 = datetime.strftime(dt2, "%Y-%m-%d")
        dt3 = datetime.today() - timedelta (days=360)
        dt3 = datetime.strftime(dt3, "%Y-%m-%d")
        price_today = YahooCrawler.read_dayprice(hist_price_stock,dt1,"-")
        price_6m_ago = YahooCrawler.read_dayprice(hist_price_stock,dt2,"+")
        price_1y_ago = YahooCrawler.read_dayprice(hist_price_stock,dt3,"+")
        change_price_6m = round(((price_today[1]-price_6m_ago[1]) / price_6m_ago[1])*100,2)
        change_price_1y = round(((price_today[1]-price_1y_ago[1]) / price_1y_ago[1])*100,2)

        #12 Dreimonatsreversal
        dt_today = datetime.today()
        m = dt_today.month
        y = dt_today.year
        d = dt_today.day
        dates = []
        for i in range(4):
            m -= 1
            if m == 0:
                y -= 1
                m = 12
            ultimo = calendar.monthrange(y,m)[1]
            dates.append(datetime.strftime(date(y,m,ultimo), "%Y-%m-%d"))
        stock_price = []
        index_price = []
        for i in dates:
            pr1 = YahooCrawler.read_dayprice(hist_price_stock, i, "-")
            stock_price.append([pr1[0],round(pr1[1],2)])
            pr2 = YahooCrawler.read_dayprice(hist_price_index, i, "-")
            index_price.append([pr2[0],round(pr2[1],2)])

        stock_change = []
        index_change = []
        for i in range(3,0,-1):
            stock_change.append(round(((stock_price[i][1] - stock_price[i-1][1]) / stock_price[i-1][1])*100,2))
            index_change.append(round(((index_price[i][1] - index_price[i-1][1]) / index_price[i-1][1])*100,2))

        if marketcap[0] < 200000000: cap = "SmallCap"
        elif marketcap[0] < 5000000000: cap = "MidCap"
        else: cap = "LargeCap"

        lm_points = 0
        lm_score = {}

        #1 - check RoE
        if roe == "N/A": lm_score["roe"] = 0
        elif roe > 20: lm_score["roe"] = 1
        elif roe < 10: lm_score["roe"] = -1
        else: lm_score["roe"] = 0

        #2 - check ebit_marge
        if fin in ["J","Y"]: lm_score["ebit_marge"] = 0
        else:
            if ebit_marge > 12 and fin.upper() == "N": lm_score["ebit_marge"] = 1
            elif ebit_marge < 6 and fin.upper() == "N": lm_score["ebit_marge"] = -1
            else: lm_score["ebit_marge"] = 0

        #3 - check eq-ratio
        if fin in ["J","Y"]:
            if eq_ratio > 10: lm_score["eq_ratio"] = 1
            elif eq_ratio < 5: lm_score["eq_ratio"] = -1
            else: lm_score["eq_ratio"] = 0
        else:
            if eq_ratio > 25: lm_score["eq_ratio"] = 1
            elif eq_ratio < 15: lm_score["eq_ratio"] = -1
            else: lm_score["eq_ratio"] = 0

        #4 - check pe-ratio
        if pe_ratio <12 and pe_ratio > 0: lm_score["pe_ratio"] = 1
        elif pe_ratio >16 or pe_ratio <0: lm_score["pe_ratio"] = -1
        else: lm_score["pe_ratio"] = 0

        #5 - check pe-ratio history
        if pe_ratio_hist <12 and pe_ratio_hist > 0: lm_score["pe_ratio_hist"] = 1
        elif pe_ratio_hist >16 or pe_ratio_hist <0: lm_score["pe_ratio_hist"] = -1
        else: lm_score["pe_ratio_hist"] = 0

        #6 - check rating
        if cap == "SmallCap":
            if rating == "N/A": lm_score["rating"] = 0
            elif rating <= 2: lm_score["rating"] = 1
            elif rating >= 4: lm_score["rating"] = -1
            else: lm_score["rating"] = 0
        else:
            if rating == "N/A": lm_score["rating"] = 0
            elif rating >= 4: lm_score["rating"] = 1
            elif rating <= 2: lm_score["rating"] = -1
            else: lm_score["rating"] = 0

        #7 - check to quarter numbers
        if reaction >1: lm_score["reaction"] = 1
        elif reaction <-1: lm_score["reaction"] = -1
        else: lm_score["reaction"] = 0

        #8 - check profit revision
        if profit_revision >1: lm_score["profit_revision"] = 1
        elif profit_revision <-1: lm_score["profit_revision"] = -1
        else: lm_score["profit_revision"] = 0

        if next_year_est_current >1: lm_score["next_year_est_current"] = 1
        elif next_year_est_current <-1: lm_score["next_year_est_current"] = -1
        else: lm_score["next_year_est_current"] = 0

        #9 - change price 6 month
        if change_price_6m >5: lm_score["change_price_6m"] = 1
        elif change_price_6m <-5: lm_score["change_price_6m"] = -1
        else: lm_score["change_price_6m"] = 0

        #10 - change price 1 year
        if change_price_1y >5: lm_score["change_price_1y"] = 1
        elif change_price_1y <-5: lm_score["change_price_1y"] = -1
        else: lm_score["change_price_1y"] = 0

        #11 - price momentum
        if lm_score["change_price_6m"] == 1 and lm_score["change_price_1y"] in [0,-1]:
            lm_score["price_momentum"] = 1
        elif lm_score["change_price_6m"] == -1 and lm_score["change_price_1y"] in [0,1]:
            lm_score["price_momentum"] = -1
        else: lm_score["price_momentum"] = 0

        #12 month reversal effect
        if cap == "LargeCap":
            if stock_change[2]<index_change[2] and stock_change[1]<index_change[1] and stock_change[0]<index_change[0]:
                lm_score["3monatsreversal"] = 1
            elif stock_change[2]>index_change[2] and stock_change[1]>index_change[1] and stock_change[0]>index_change[0]:
                lm_score["3monatsreversal"] = -1
            else: lm_score["3monatsreversal"] = 0
        else:
            lm_score["3monatsreversal"] = 0

        #13 - profit growth
        if profit_growth >5: lm_score["profit_growth"] = 1
        elif profit_growth <-5: lm_score["profit_growth"] = -1
        else: lm_score["profit_growth"] = 0

        # print format marketcap
        print_cap = YahooCrawler.print_num_abbr(marketcap[0])

        lm_sum = 0
        for val in lm_score.values(): lm_sum += val
        # overall recomendation levermann full
        if cap in ["SmallCap","MidCap"]:
            if lm_sum >=7: rec = "Possible Buy"
            elif lm_sum in [5,6]: rec = "Possible Holding"
            else: rec = "Possible Sell"
        else:
            if lm_sum >=4: rec = "Possible Buy"
            elif lm_sum in [3]: rec = "Possible Holding"
            else: rec = "Possible Sell"

        lm_sum_light = lm_score["roe"] + lm_score["ebit_marge"] + lm_score["pe_ratio"] \
                       + lm_score["reaction"] + lm_score["change_price_6m"]
        # overall recomendation levermann full
        if cap in ["SmallCap","MidCap"]:
            if lm_sum_light >=4: rec_light = "Possible Buy"
            else: rec_light = "Possible Sell"
        else:
            if lm_sum_light >=3: rec_light = "Possible Buy"
            else: rec_light = "Possible Sell"

        output = []
        output.append(["Name",name,"","","","","","","","","Details"])
        output.append(["Ticker",stock,"","","","","","","","",1,"Return on Equity (ttm)",stat1["Return on Equity (ttm)"]])
        output.append(["Index", index,"","","","","","","","",2,"EBIT (ttm)",str(YahooCrawler.print_num_abbr(insstat["EBIT"][0])),])
        output.append(["MCap", print_cap,"","","","","","","","","","Revenue (ttm)",str(YahooCrawler.print_num_abbr(insstat["Total Revenue"][0]))])
        if ebit_marge == "N/A": output.append(["","","","","","","","","","","","EBIT Marge (ttm)","N/A"])
        else: output.append(["","","","","","","","","","","","EBIT Marge (ttm)",str(ebit_marge)+"%"])
        output.append(["Nr.", "Levermann Checkliste FULL","",1,-1])
        output.append([1,"Eigenkapitalrendite RoE","Return on Equity RoE",">20","<10",roe,lm_score["roe"]])
        output.append([2,"EBIT-Marge","EBIT-Margin",">12","<6",ebit_marge,lm_score["ebit_marge"]])
        output.append([3,"Eigenkapitalquote","Equity Ratio",">25","<15",eq_ratio,lm_score["eq_ratio"]])
        output.append([4,"KGV Aktuell","P/E-Ratio History","<12",">16",pe_ratio_hist,lm_score["pe_ratio_hist"]])
        output.append([5,"KGV 5 Jahre Mittel","P/E-Ratio Actual","<12",">16",pe_ratio,lm_score["pe_ratio"]])
        output.append([6,"Analystenmeinung","Analyst Opinions","<2",">4",rating,lm_score["rating"]])
        output.append([7,"Reaktion auf Quartalszahlen","Reaction to quarter numbers",">1","<-1",reaction,lm_score["reaction"]])
        output.append([8,"Gewinnrevision","Profit Revision",">5","<5",round (profit_revision, 2),lm_score["profit_revision"]])
        output.append([9,"Kurs Heute vs. Kurs 6M","Price Change for 6 month",">5","<5",change_price_6m,lm_score["change_price_6m"]])
        output.append([10,"Kurs Heute vs. Kurs 1J","Price Change for 12 month","","",change_price_1y,lm_score["change_price_1y"]])
        output.append([11,"Kursmomentum Steigend","Price Momentum","","","",lm_score["price_momentum"]])
        output.append([12,"Dreimonatsreversal","3 Month Reversal Effect","","","",lm_score["3monatsreversal"]])
        output.append([13,"Gewinnwachstum","Profit Growth",">5","<5",profit_growth,lm_score["profit_growth"]])
        output.append(["",rec,rec,"","","",lm_sum])
        output.append(["",cap,cap])
        if fin.upper() in ["Y","J"]:
            output.append (["","Finanzwert","Financial Stock"])
        else:
            output.append (["","Kein Finanzwert","No Financial Stock"])

        for i in range(3): output.append([""])
        output.append (["Nr.", "Levermann Checkliste LIGHT", "", 1, -1])
        output.append([1,"Eigenkapitalrendite RoE","Return on Equity RoE",">20","<10",roe,lm_score["roe"]])
        output.append([2,"EBIT-Marge","EBIT-Margin",">12","<6",ebit_marge,lm_score["ebit_marge"]])
        output.append([3,"KGV 5 Jahre Mittel","P/E-Ratio Actual","<12",">16",pe_ratio,lm_score["pe_ratio"]])
        output.append([4,"Reaktion auf Quartalszahlen","Reaction to quarter numbers",">1","<-1",reaction,lm_score["reaction"]])
        output.append([5,"Kurs Heute vs. Kurs 6M","Price Change for 6 month",">5","<5",change_price_6m,lm_score["change_price_6m"]])
        output.append (["", rec_light, rec_light, "", "", "", lm_sum_light])
        output.append (["", cap, cap])
        if fin.upper() in ["Y","J"]:
            output.append (["","Finanzwert","Financial Stock"])
        else:
            output.append (["","Kein Finanzwert","No Financial Stock"])

        stop = timeit.default_timer()
        print ("Total time working on stock",stock,":",round (stop - start, 0),"sec...")

        # delete input stocks in input-xlsx
        if ws_db["G2"].value.upper() in ["J","Y"]:
            cell_list = ["","",""]
            i = str (idx_stock + 3)
            ws_db.range ("A" + i + ":" + "C" + i).value = cell_list

        save_xls (stock, output, out)

    wb.save(fn)













