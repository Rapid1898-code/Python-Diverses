from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

wb = load_workbook(("Test.xlsx"))
ws = wb["TAB"]

print(ws["A1"].value)
ws["A1"] = 97
print(ws["A1"].value)
print(ws.cell(row=1, column=1).value)

ws.insert_cols(2,1)
wb.save("Test.xlsx")
